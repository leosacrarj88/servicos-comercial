# -*- coding: utf-8 -*-
"""
Mapa Comercial por categorias
✅ Um único arquivo .py
✅ Auto-start Streamlit ao rodar pelo Play do VS Code (sem loop)
✅ Resultados persistentes (st.form + st.session_state)
✅ Busca por categorias:
   - Google Places Nearby Search + Details (telefone, website, horários, url) se tiver API key
   - OSM/Overpass fallback (cache + fallback de endpoints)
✅ Cards primeiro, em grade (até 5 por linha, configurável na sidebar)
✅ Mapa interativo (Folium/OSM) com MarkerCluster + links Google Maps/Rotas

Requisitos:
pip install streamlit pandas requests geopy folium streamlit-folium
"""

# ===============================
# EXECUÇÃO
# ===============================
# ===============================
# IMPORTS
# ===============================
import time
import math
import json
import re
import html
import base64
import requests
import pandas as pd
import streamlit as st

# ===============================
# SECRETS (Safe access)
# - No Streamlit Cloud, st.secrets existe.
# - No local, pode não existir secrets.toml e o Streamlit levanta StreamlitSecretNotFoundError.
# ===============================
def _safe_secrets_get(key: str, default=None):
    try:
        return st.secrets.get(key, default)  # type: ignore[attr-defined]
    except Exception:
        return default


def _safe_secrets_has(key: str) -> bool:
    try:
        return key in st.secrets  # type: ignore[attr-defined]
    except Exception:
        return False
from geopy.geocoders import Nominatim
from geopy.exc import GeocoderTimedOut
import folium
from folium.plugins import MarkerCluster
from streamlit_folium import st_folium
import sys
import os
import io
import hashlib

# ===============================
# GOOGLE SHEETS - PADRÕES
# ===============================
DEFAULT_GSHEET_ID = (os.getenv("GSHEET_ID", "106O5MwhhB9LV55tXnJTq4kObNNt0LU4jQmEsxGDCrlg") or "106O5MwhhB9LV55tXnJTq4kObNNt0LU4jQmEsxGDCrlg").strip()
DEFAULT_GSHEET_TAB = (os.getenv("GSHEET_TAB", "Clientes") or "Clientes").strip()
_GS_FIXED_HEADERS = ["ID_CLIENTE", "Segmento"]
_GS_DYNAMIC_HEADERS = ["Segmento_TXT", "Executiva Pixel", "Empresa (Cliente)", "Responsável pela Empresa", "Telefone", "E-mail", "Já fiz contato?", "Data de contato", "Observações", "Site", "Endereço", "Bairro", "CEP", "Atualizado em", "Foto", "Foto_AppSheet"]
_GS_EXPORT_HEADERS = _GS_FIXED_HEADERS + _GS_DYNAMIC_HEADERS



# ===============================
# FILTROS DE LOCALIZAÇÃO (Sugestões de endereço)
# ===============================
# Se vazio/None, aceita qualquer UF.
# Exemplo para restringir: ALLOWED_UF = {"RJ", "SP"}
ALLOWED_UF = set()  # type: set[str]

import subprocess
import socket
from pathlib import Path


from io import BytesIO
from datetime import datetime
try:
    from zoneinfo import ZoneInfo  # py3.9+
except Exception:
    ZoneInfo = None

try:
    from PIL import Image as PILImage
except Exception:
    PILImage = None

from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage

try:
    import gspread
    from google.oauth2.service_account import Credentials
    from googleapiclient.discovery import build
    from googleapiclient.http import MediaIoBaseUpload
except Exception:
    gspread = None
    Credentials = None

# (ngrok removido: não necessário para este app)

# ===============================
# APP
# ===============================


# ===============================
# HISTÓRICO DE ENDEREÇOS (JSON)
# - Salva/Carrega um arquivo JSON na mesma pasta do app/exec
# - Mantém uma lista de endereços recentes para reuso
# ===============================

def _app_base_dir() -> str:
    """Pasta onde o app deve salvar arquivos (mesma pasta do .py ou do executável)."""
    try:
        if getattr(sys, "frozen", False):  # PyInstaller / exe
            return os.path.dirname(sys.executable)
        return os.path.dirname(os.path.abspath(__file__))
    except Exception:
        # fallback: pasta atual do processo
        return os.getcwd()


def _addr_history_path() -> str:
    return os.path.join(_app_base_dir(), "historico_enderecos.json")


def load_address_history(limit: int = 25) -> list[str]:
    """Retorna lista de endereços (mais recentes primeiro)."""
    fp = _addr_history_path()
    try:
        if not os.path.exists(fp):
            return []
        data = json.loads(Path(fp).read_text(encoding="utf-8"))
        if isinstance(data, dict):
            data = data.get("enderecos", [])
        if not isinstance(data, list):
            return []
        out = []
        for x in data:
            if isinstance(x, str):
                s = x.strip()
                if s:
                    out.append(s)
        # remove duplicados mantendo ordem
        seen = set()
        uniq = []
        for s in out:
            k = s.lower()
            if k not in seen:
                uniq.append(s)
                seen.add(k)
        return uniq[: max(1, limit)]
    except Exception:
        return []


def save_address_history(addresses: list[str]) -> None:
    fp = _addr_history_path()
    try:
        payload = {"enderecos": addresses}
        Path(fp).write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    except Exception:
        # não pode quebrar o app por falha de IO
        return


def add_address_to_history(addr: str, limit: int = 25) -> None:
    a = (addr or "").strip()
    if not a:
        return
    hist = load_address_history(limit=limit)
    # coloca no topo, removendo duplicados
    new_hist = [a] + [x for x in hist if x.strip().lower() != a.lower()]
    save_address_history(new_hist[: max(1, limit)])



# ===============================
# EXPORTAÇÃO PARA EXCEL (Incremental)
# ===============================
EXPORT_TEMPLATE_FILENAME = "Prospecção Vanessa Pixel Rio.xlsx"
DEFAULT_EXPORT_SHEET_NAME = "Moinho clientes"

_EXPORT_REQUIRED_HEADERS = ["Segmento", "Cliente", "Responsável", "Contato", "Já fiz contato?", "Data de contato", "E-mail", "Endereço", "CEP", "Atualizado em", "Executiva", "Foto"]
def _norm_header(v) -> str:
    if v is None:
        return ""
    return str(v).strip().lower()


def _first_empty_header_col(ws, header_row: int, start_col: int = 1, end_col: int = 20):
    """
    Encontra a 1ª coluna (no header_row) cujo header está vazio.
    Importante: evita criar colunas novas lá no fim da planilha (muitas têm formatação até colunas bem à direita).
    """
    for c in range(start_col, end_col + 1):
        v = ws.cell(header_row, c).value
        if v is None or str(v).strip() == "":
            return c
    return None


def _find_header_row(ws, scan_rows=30, scan_cols=60):
    # considera "achou" se pelo menos Segmento e Cliente estão na mesma linha
    for r in range(1, scan_rows + 1):
        row_vals = [_norm_header(ws.cell(r, c).value) for c in range(1, scan_cols + 1)]
        if ("segmento" in row_vals) and ("cliente" in row_vals):
            # guarda todas as colunas (inclusive duplicadas) por header
            all_cols = {}
            for c, v in enumerate(row_vals, start=1):
                if v:
                    all_cols.setdefault(v, []).append(c)
            return r, all_cols
    return None, None


def _ensure_export_layout(ws):
    """
    Garante que a aba tenha os headers:
    Segmento | Executiva | Endereço origem | Cliente | Endereço | Responsável | Contato | Atualizado em
    E tenta manter tudo em colunas visíveis (A..T), mesmo que a planilha tenha
    formatação até colunas muito à direita.
    """
    header_row, all_cols = _find_header_row(ws)

    # Se não achou header: cria do zero
    if header_row is None:
        header_row = 1
        for i, h in enumerate(_EXPORT_REQUIRED_HEADERS, start=1):
            ws.cell(header_row, i).value = h
        all_cols = {_norm_header(h): [i] for i, h in enumerate(_EXPORT_REQUIRED_HEADERS, start=1)}

    # Layout específico da sua planilha anexa (aba "Moinho clientes"):
    # Coluna 2 vem "Responsável" mas na prática é "Executiva" (valor "Vanessa")
    # Coluna 4 vem "Responsável " (com espaço).
    v2 = _norm_header(ws.cell(header_row, 2).value)
    v4 = _norm_header(ws.cell(header_row, 4).value)
    if v2 in ("responsável", "responsavel") and v4 in ("responsável", "responsavel"):
        ws.cell(header_row, 2).value = "Executiva"
        ws.cell(header_row, 4).value = "Responsável"
        # atualiza all_cols de forma consistente
        all_cols.setdefault("executiva", [2])
        all_cols["executiva"] = [2]
        all_cols.setdefault("responsável", [4])
        all_cols["responsável"] = [4]
        all_cols["responsavel"] = [4]

        # Se a planilha estiver no formato antigo (sem Endereço origem), insere 1 coluna logo após "Executiva"
    # para manter "Endereço origem" AO LADO de "Executiva" (como você pediu).
    try:
        has_exec = "executiva" in (all_cols or {})
        has_origin = ("endereço origem" in (all_cols or {})) or ("endereco origem" in (all_cols or {}))
        if has_exec and (not has_origin):
            exec_col = (all_cols.get("executiva") or [None])[0]
            if isinstance(exec_col, int) and exec_col > 0:
                target_col = int(exec_col) + 1
                ws.insert_cols(target_col, 1)
                ws.cell(header_row, target_col).value = "Endereço origem"

                # Reconstroi o mapa de headers após o insert
                all_cols = {}
                for c in range(1, 61):
                    v = _norm_header(ws.cell(header_row, c).value)
                    if v:
                        all_cols.setdefault(v, []).append(c)
    except Exception:
        pass

# Se a planilha ainda estiver no formato antigo (sem Endereço), insere uma coluna logo após "Cliente"
    # para garantir que "Endereço" fique AO LADO de "Cliente" (como você pediu).
    try:
        has_cliente = "cliente" in (all_cols or {})
        has_end = ("endereço" in (all_cols or {})) or ("endereco" in (all_cols or {}))
        if has_cliente and (not has_end):
            cliente_col = (all_cols.get("cliente") or [None])[0]
            if isinstance(cliente_col, int) and cliente_col > 0:
                target_col = int(cliente_col) + 1
                ws.insert_cols(target_col, 1)
                ws.cell(header_row, target_col).value = "Endereço"

                # Reconstroi o mapa de headers após o insert (porque as colunas mudaram)
                all_cols = {}
                for c in range(1, 61):
                    v = _norm_header(ws.cell(header_row, c).value)
                    if v:
                        all_cols.setdefault(v, []).append(c)
    except Exception:
        pass

    # Monta mapa final header->col (preferindo colunas mais à esquerda)
    def _pick_col(h_norm: str):
        cols = (all_cols or {}).get(h_norm) or []
        return cols[0] if cols else None

    # "Responsável" pode vir com ou sem acento
    resp_col = _pick_col("responsável") or _pick_col("responsavel")

    cols = {
        "Segmento": _pick_col("segmento"),
        "Executiva": _pick_col("executiva"),
        "Endereço origem": _pick_col("endereço origem") or _pick_col("endereco origem"),
        "Cliente": _pick_col("cliente"),
        "Endereço": _pick_col("endereço") or _pick_col("endereco"),
        "Responsável": resp_col,
        "Contato": _pick_col("contato"),
        "Atualizado em": _pick_col("atualizado em"),
    }

    # Garante headers faltantes usando primeira coluna vazia (A..T) antes de "jogar" para o fim
    def _ensure_header(title: str):
        hn = _norm_header(title)
        if hn in ("responsável", "responsavel"):
            if cols["Responsável"] is not None:
                return cols["Responsável"]
        if hn == "executiva" and cols["Executiva"] is not None:
            return cols["Executiva"]
        if hn in ("endereço origem", "endereco origem") and cols.get("Endereço origem") is not None:
            return cols["Endereço origem"]
        if hn == "segmento" and cols["Segmento"] is not None:
            return cols["Segmento"]
        if hn == "cliente" and cols["Cliente"] is not None:
            return cols["Cliente"]
        if hn in ("endereço", "endereco") and cols.get("Endereço") is not None:
            return cols["Endereço"]
        if hn == "contato" and cols["Contato"] is not None:
            return cols["Contato"]
        if hn == "atualizado em" and cols["Atualizado em"] is not None:
            return cols["Atualizado em"]

        c = _first_empty_header_col(ws, header_row, 1, 20)
        if c is None:
            # fallback: após o último header não-vazio até col 60; senão, max_column+1
            last_h = 0
            for cc in range(1, 61):
                vv = ws.cell(header_row, cc).value
                if vv is not None and str(vv).strip() != "":
                    last_h = cc
            c = last_h + 1 if last_h else (ws.max_column or 1) + 1

        ws.cell(header_row, c).value = title
        return c

    # garante todos
    cols["Segmento"] = cols["Segmento"] or _ensure_header("Segmento")
    cols["Executiva"] = cols["Executiva"] or _ensure_header("Executiva")
    cols["Endereço origem"] = cols.get("Endereço origem") or _ensure_header("Endereço origem")
    cols["Cliente"] = cols["Cliente"] or _ensure_header("Cliente")
    cols["Endereço"] = cols.get("Endereço") or _ensure_header("Endereço")
    cols["Responsável"] = cols["Responsável"] or _ensure_header("Responsável")
    cols["Contato"] = cols["Contato"] or _ensure_header("Contato")
    cols["Atualizado em"] = cols["Atualizado em"] or _ensure_header("Atualizado em")

    return header_row, cols


def _last_filled_row_any(ws, header_row: int, cols_to_check):
    """
    Última linha com qualquer valor em qualquer uma das colunas informadas.
    Isso evita 'start_row' errar quando a planilha tem várias linhas com Segmento/Executiva já preenchidos.
    """
    cols_to_check = [c for c in (cols_to_check or []) if isinstance(c, int) and c > 0]
    if not cols_to_check:
        return header_row

    max_r = ws.max_row or header_row
    for r in range(max_r, header_row, -1):
        for c in cols_to_check:
            v = ws.cell(r, c).value
            if v is not None and str(v).strip() != "":
                return r
    return header_row


def _digits_only(s: str) -> str:
    return "".join(ch for ch in (s or "") if ch.isdigit())



def _clean_text(x) -> str:
    if x is None:
        return ""
    s = str(x).strip()
    if s.lower() == "nan":
        return ""
    return s


def _split_multi(value: str):
    s = _clean_text(value)
    if not s:
        return []
    # Split by common separators
    parts = re.split(r"[;|,/]+|\s/\s", s)
    return [p.strip() for p in parts if p and p.strip()]


def _normalize_phone_br(raw: str) -> str:
    s = _clean_text(raw)
    if not s:
        return ""
    # Keep digits and plus
    digits = re.sub(r"\D", "", s)
    if not digits:
        return ""
    # If already has country code 55
    if digits.startswith("55") and len(digits) >= 12:
        return "+" + digits
    # If BR local with DDD (10/11 digits)
    if len(digits) in (10, 11):
        return "+55" + digits
    # fallback
    return "+" + digits if not s.startswith("+") else s


def _get_primary_phone_ddd(row) -> str:
    """
    Retorna SOMENTE o telefone no formato DDD+NÚMERO (apenas dígitos).
    - Se vier com +55, remove o 55.
    - Se vier com DDD (10/11 dígitos), mantém.
    - Se vier 0800/sem DDD, retorna os dígitos como estão.
    """
    raw_tel = _clean_text(row.get("Telefone") if hasattr(row, "get") else "")
    if not raw_tel:
        return ""

    for p in _split_multi(raw_tel):
        digits = re.sub(r"\D", "", str(p))
        if not digits:
            continue

        # remove país 55 quando estiver presente
        if digits.startswith("55") and len(digits) in (12, 13, 14):
            digits = digits[2:]

        # se tiver 10/11 dígitos, é DDD+numero
        if len(digits) in (10, 11):
            return digits

        # 0800 ou outros formatos sem DDD
        if digits.startswith("0800") and len(digits) >= 10:
            return digits

        return digits

    return ""


def _get_email_or_site(row) -> str:
    """
    Preferência:
    1) e-mail (se existir)
    2) site (se não tiver e-mail)
    """
    email = _get_email_row(row)
    if email:
        return email

    site = _clean_text(row.get("Website") if hasattr(row, "get") else "")
    if not site or site == "-":
        for k in ["Site", "URL", "Url", "url", "Website"]:
            site = _clean_text(row.get(k) if hasattr(row, "get") else "")
            if site and site != "-":
                break
    return site if site and site != "-" else ""


def _get_site_row(row) -> str:
    """Retorna somente o site/URL da empresa, sem misturar com e-mail."""
    for key in ["Website", "Site", "URL", "Url", "url", "Maps"]:
        v = _clean_text(row.get(key) if hasattr(row, "get") else "")
        if v and v != "-":
            if _extract_email_from_text(v):
                continue
            return v
    return ""


def _extract_cep(text: str) -> str:
    s = _clean_text(text)
    if not s:
        return ""
    m = re.search(r"\b\d{5}-\d{3}\b", s)
    if m:
        return m.group(0)
    m = re.search(r"\b\d{8}\b", s)
    if m:
        raw = m.group(0)
        return f"{raw[:5]}-{raw[5:]}"
    return ""


def _clean_endereco_full(s: str) -> str:
    """
    Mantém o endereço COMPLETO (incluindo complementos),
    apenas removendo ruídos comuns do final como ", Brasil".
    """
    s = _clean_text(s)
    if not s:
        return ""
    s = re.sub(r",\s*Brasil\s*$", "", s, flags=re.IGNORECASE).strip()
    return s



def _split_endereco_brasil(full: str):
    """
    Heurística para Google formatted_address / strings de endereço no Brasil.
    Retorna: (endereco_rua_num, bairro, cidade, cep)

    Casos comuns do Google:
      "R. Teixeira de Melo, 31 - Ipanema, Rio de Janeiro - RJ, 22410-001, Brasil"
      "Av. Brasil, 5000 - Bonsucesso, Rio de Janeiro - RJ, 21040-360, Brasil"
    """
    s = _clean_text(full)
    if not s or s.lower().startswith("endereço não"):
        return (s, "", "", "")

    cep = _extract_cep(s)

    parts = [p.strip() for p in s.split(",") if p and p.strip()]
    parts = [p for p in parts if p.lower() not in ("brasil",)]

    endereco = ""
    bairro = ""
    cidade = ""

    if not parts:
        return (s, "", "", cep)

    # 1) Endereço base (rua/av)
    first = parts[0].strip()
    # Caso: "Rua X, 123 - Bairro" vem inteiro no first
    if " - " in first:
        left, right = first.split(" - ", 1)
        endereco = left.strip()
        # se o lado esquerdo já tem número, ótimo; se não, vamos tentar adicionar depois
        if right and not re.search(r"\s-\s*[A-Z]{2}\b", right):
            bairro = right.strip()
    else:
        endereco = first

    # 2) Detecta cidade (preferência: "Cidade - UF")
    city_idx = None
    for i, p in enumerate(parts):
        if re.search(r"\s-\s*[A-Z]{2}\b", p):
            cidade = re.sub(r"\s-\s*[A-Z]{2}\b.*$", "", p).strip()
            city_idx = i
            break

    # 3) Se ainda não achou cidade: pega o último item "útil" (antes do CEP)
    if not cidade:
        for p in reversed(parts):
            if cep and cep in p:
                continue
            if re.fullmatch(r"\d{5}-\d{3}", p) or re.fullmatch(r"\d{8}", p):
                continue
            if p.lower() == "brasil":
                continue
            cidade = p.strip()
            city_idx = parts.index(p)
            break

    # 4) Número + Bairro costumam vir no 2º item: "31 - Ipanema" ou "31" ou "31 - Bairro"
    def _apply_num_bairro(segment: str):
        nonlocal endereco, bairro
        seg = segment.strip()
        mnb = re.match(r"^(\d+)\s*-\s*(.+)$", seg)
        if mnb:
            num = mnb.group(1).strip()
            rest = mnb.group(2).strip()
            # quando vier "num - complemento - bairro", pega o ÚLTIMO como bairro
            b = rest.split(" - ")[-1].strip() if " - " in rest else rest
            if num and num not in endereco:
                endereco = f"{endereco}, {num}".strip().strip(",")
            if not bairro and b:
                bairro = b
            return True
        if re.fullmatch(r"\d+", seg):
            num = seg
            if num and num not in endereco:
                endereco = f"{endereco}, {num}".strip().strip(",")
            return True
        return False

    if len(parts) >= 2:
        # se parts[1] é número/bairro, aplica
        _apply_num_bairro(parts[1])

    # 5) Se ainda não tiver bairro, tenta pegar o item imediatamente antes da cidade
    if not bairro and city_idx is not None and city_idx - 1 >= 1:
        cand = parts[city_idx - 1].strip()
        # pode ser "31 - Ipanema" ou só "Ipanema"
        if not _apply_num_bairro(cand):
            if cand and not re.search(r"\s-\s*[A-Z]{2}\b", cand) and (not cep or cep not in cand):
                # evita pegar o próprio endereço
                if cand != endereco:
                    bairro = cand

    # 6) Caso ainda esteja vazio e exista parts[2], use como bairro (quando não for cidade)
    if not bairro and len(parts) >= 3:
        cand = parts[2].strip()
        if cand and not re.search(r"\s-\s*[A-Z]{2}\b", cand) and (not cep or cep not in cand):
            # evita repetir cidade
            if cand != cidade:
                bairro = cand

    # 7) Limpeza final
    if cep:
        cidade = cidade.replace(cep, "").strip(" ,")
        bairro = bairro.replace(cep, "").strip(" ,")

    return (endereco, bairro, cidade, cep)



def _get_endereco_fields(row):
    """
    Retorna (endereco_completo, bairro, cidade, cep).

    - Endereço: SEMPRE completo (complementos) quando existir na fonte (Google formatted_address).
    - Bairro: somente o bairro.
    """
    end_full_raw = _clean_text(row.get("Endereço") if hasattr(row, "get") else "")
    endereco_full = _clean_endereco_full(end_full_raw)

    # Se o OSM tiver só rua/num, ainda assim usamos isso como "Endereço"
    if not endereco_full:
        endereco_full = _clean_text(row.get("Endereco") if hasattr(row, "get") else "")

    bairro = _clean_text(row.get("Bairro") if hasattr(row, "get") else "")
    cidade = _clean_text(row.get("Cidade") if hasattr(row, "get") else "")
    cep = _clean_text(row.get("CEP") if hasattr(row, "get") else "") or _clean_text(row.get("Cep") if hasattr(row, "get") else "")

    # Se não veio estruturado, tenta inferir bairro/cidade/cep do endereço completo
    if endereco_full and not (bairro or cidade or cep):
        _e, b, c, z = _split_endereco_brasil(endereco_full)
        bairro = bairro or b
        cidade = cidade or c
        cep = cep or z

    # CEP fallback
    if not cep and endereco_full:
        cep = _extract_cep(endereco_full)

    return endereco_full, bairro, cidade, cep






# ===============================
# FOTO (1 por empresa) para Google Sheets
# - Preferência: Street View (fachada) quando houver cobertura
# - Fallback: Places Photo (melhor foto disponível)
# ===============================

@st.cache_data(ttl=24*3600)
def _streetview_has_pano(lat: float, lon: float, api_key: str) -> bool:
    if not api_key:
        return False
    try:
        url = "https://maps.googleapis.com/maps/api/streetview/metadata"
        params = {"location": f"{lat},{lon}", "radius": 80, "source": "outdoor", "key": api_key}
        r = requests.get(url, params=params, timeout=15)
        j = r.json() if r.ok else {}
        return (j.get("status") == "OK")
    except Exception:
        return False


def _streetview_static_url(lat: float, lon: float, api_key: str, size: str = "640x360") -> str:
    # heading/pitch neutros; prioriza outdoor
    return (
        "https://maps.googleapis.com/maps/api/streetview"
        f"?size={size}&location={lat},{lon}&radius=80&source=outdoor&fov=90&pitch=0&key={api_key}"
    )


def _places_photo_url(photo_reference: str, api_key: str, maxwidth: int = 1200) -> str:
    return (
        "https://maps.googleapis.com/maps/api/place/photo"
        f"?maxwidth={int(maxwidth)}&photo_reference={photo_reference}&key={api_key}"
    )


from functools import lru_cache

@lru_cache(maxsize=4096)
def _resolve_redirect_url(url: str) -> str:
    """Resolve redirects (302) and return final URL.

    Motivação: o Google Sheets (função IMAGE) pode não seguir redirect/URL com key,
    e chaves com restrição de referrer podem falhar. Ao resolver o redirect do
    endpoint /place/photo, obtemos uma URL final (googleusercontent) que costuma
    funcionar bem no Sheets.
    """
    try:
        r = requests.get(url, allow_redirects=False, timeout=15)
        # Places Photo normalmente retorna 302 com header Location
        loc = r.headers.get("Location") or r.headers.get("location")
        if loc:
            return loc
    except Exception:
        pass
    return url



def _best_company_photo_url(row, api_key: str) -> str:
    """Retorna UMA URL de imagem para exportar no Google Sheets.

    Estratégia (mais compatível com =IMAGE no Sheets):
    1) Tenta Places Photo (PhotoRef) e resolve redirect para URL final (googleusercontent).
       -> normalmente funciona melhor no Sheets do que a URL com key (maps.googleapis.com).
    2) Fallback: Street View Static (fachada), se houver panorama.
    """
    if not api_key:
        return ""

    # 1) Places Photo (preferido para Sheets)
    try:
        pref = (row.get("PhotoRef") if hasattr(row, "get") else "") or ""
        pref = str(pref).strip()
        if pref:
            url = _places_photo_url(pref, api_key, maxwidth=1200)
            return _resolve_redirect_url(url)
    except Exception:
        pass

    # 2) Street View (fachada) - fallback
    try:
        lat = row.get("Latitude") if hasattr(row, "get") else None
        lon = row.get("Longitude") if hasattr(row, "get") else None
        lat = float(lat) if lat not in (None, "", "nan") else None
        lon = float(lon) if lon not in (None, "", "nan") else None
    except Exception:
        lat, lon = None, None

    if lat is not None and lon is not None and _streetview_has_pano(lat, lon, api_key):
        return _streetview_static_url(lat, lon, api_key)

    return ""

def _extract_email_from_text(s: str) -> str:
    s = _clean_text(s)
    if not s:
        return ""
    m = re.search(r"([A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,})", s)
    return m.group(1) if m else ""


def _get_email_row(row) -> str:
    # Prefer explicit email fields (OSM may provide)
    for key in ["E-mail", "Email", "email", "contact:email"]:
        v = row.get(key) if hasattr(row, "get") else None
        e = _extract_email_from_text(v)
        if e:
            return e
    # Sometimes comes embedded in website or notes
    for key in ["Website", "Site", "URL", "Url", "url", "Maps"]:
        v = row.get(key) if hasattr(row, "get") else None
        e = _extract_email_from_text(v)
        if e:
            return e
    return ""


def _build_contato(row, email: str = "") -> str:
    """
    Contato robusto e seguro para Google Sheets (sem virar fórmula):
    Prioridade: E-mail + Telefone.
    - Nunca começa com "+" (pra não virar fórmula quando value_input_option=USER_ENTERED).
    - Usa separador "; " (evita " | ").
    """
    # 1) e-mail (se já foi calculado, aproveita)
    email_final = _clean_text(email) or _get_email_row(row)

    # 2) telefones (normalizados)
    phones = []
    raw_tel = _clean_text(row.get("Telefone") if hasattr(row, "get") else "")
    for p in _split_multi(raw_tel):
        ph = _normalize_phone_br(p)
        if ph and ph not in phones:
            phones.append(ph)

    # 3) outros canais (quando existirem)
    whatsapp = _clean_text(row.get("WhatsApp") if hasattr(row, "get") else "")
    instagram = _clean_text(row.get("Instagram") if hasattr(row, "get") else "")
    facebook = _clean_text(row.get("Facebook") if hasattr(row, "get") else "")
    linkedin = _clean_text(row.get("LinkedIn") if hasattr(row, "get") else "")
    site = _clean_text(row.get("Website") if hasattr(row, "get") else "")
    maps_url = _clean_text(row.get("Maps") if hasattr(row, "get") else "")

    parts = []

    # Prioridade: email e telefone
    if email_final:
        parts.append(f"Email: {email_final}")
    if phones:
        parts.append(f"Tel: {' / '.join(phones)}")

    # Se ainda não tiver email/telefone, tenta site e maps como fallback
    if not parts:
        if site and site != "-":
            parts.append(f"Site: {site}")
        elif maps_url:
            parts.append(f"Maps: {maps_url}")

    # Complementos úteis (sem atrapalhar)
    if whatsapp:
        parts.append(f"WhatsApp: {whatsapp}")
    if instagram:
        parts.append(f"Instagram: {instagram}")
    if facebook:
        parts.append(f"Facebook: {facebook}")
    if linkedin:
        parts.append(f"LinkedIn: {linkedin}")

    # Importante: separador seguro
    return "; ".join([p for p in parts if p]).strip()



def _get_responsavel_row(row) -> str:
    # Sem scraping: pega somente se vier explícito (OSM contact:name/person)
    for key in ["Responsável", "Responsavel", "contact:name", "contact:person"]:
        v = _clean_text(row.get(key) if hasattr(row, "get") else "")
        if v:
            return v
    return ""


def _dedup_key(cliente: str, contato: str, email: str, endereco: str):
    c = _clean_text(cliente).lower()
    e = _clean_text(email).lower()
    a = _clean_text(endereco).lower()
    # phone digits priority
    digits = re.sub(r"\D", "", _clean_text(contato))
    if digits:
        return (c, f"tel:{digits}")
    if e:
        return (c, f"email:{e}")
    if a:
        return (c, f"addr:{a}")
    return (c, "")


def _pick_row_value(row, candidates):
    for c in candidates:
        if c in row and row.get(c) is not None and str(row.get(c)).strip().lower() != "nan":
            return row.get(c)
    return ""



def fetch_place_photo_bytes_export(photo_reference: str, api_key: str, maxwidth: int = 520):
    """Versão top-level da lógica de foto usada na tela, para exportações."""
    if not photo_reference or not api_key:
        return None
    url = "https://maps.googleapis.com/maps/api/place/photo"
    params = {"maxwidth": int(maxwidth), "photo_reference": photo_reference, "key": api_key}
    try:
        r = requests.get(url, params=params, timeout=20, allow_redirects=True)
        if r.status_code == 200 and r.content:
            return r.content
    except Exception:
        return None
    return None


def _get_photo_bytes_for_export(row, api_key_for_photo: str):
    """Retorna (img_bytes, mime_type) usando a mesma prioridade visual da tela."""
    try:
        api_key_for_photo = (api_key_for_photo or "").strip()
        if not api_key_for_photo:
            return (None, None)

        photo_ref = str((row.get("PhotoRef") if hasattr(row, "get") else "") or "").strip()
        if photo_ref:
            img_bytes = fetch_place_photo_bytes_export(photo_ref, api_key_for_photo, maxwidth=1200)
            if img_bytes:
                return (img_bytes, "image/jpeg")

        foto_url = _best_company_photo_url(row, api_key=api_key_for_photo)
        if foto_url:
            return _download_image_bytes(foto_url)
    except Exception:
        pass
    return (None, None)




def _photo_public_url_for_row(row, api_key_for_photo: str) -> str:
    """Retorna a URL pública direta da foto para uso no AppSheet.

    Prioridade:
    1) URL final resolvida do Places Photo (googleusercontent), mais compatível com AppSheet.
    2) Fallback para a melhor URL pública disponível.
    """
    try:
        api_key_for_photo = (api_key_for_photo or "").strip()
        if not api_key_for_photo:
            return ""

        if "photo_public_url_cache" not in st.session_state:
            st.session_state.photo_public_url_cache = {}
        cache: dict = st.session_state.photo_public_url_cache

        place_id = ""
        if hasattr(row, "get"):
            place_id = str(row.get("PlaceId") or row.get("place_id") or "").strip()
        nome = str(row.get("Nome") if hasattr(row, "get") else "")[:80].strip() or "empresa"
        endereco_key = str(row.get("Endereço") if hasattr(row, "get") else "")
        key = place_id or hashlib.md5((nome + endereco_key).encode("utf-8", errors="ignore")).hexdigest()

        if key in cache:
            return cache[key] or ""

        public_url = ""
        photo_ref = str((row.get("PhotoRef") if hasattr(row, "get") else "") or "").strip()
        if photo_ref:
            try:
                direct_url = _resolve_redirect_url(_places_photo_url(photo_ref, api_key_for_photo, maxwidth=1600))
                if direct_url and any(x in direct_url for x in ("googleusercontent.com", "ggpht.com", "googleapis.com")):
                    public_url = direct_url
            except Exception:
                pass

        if not public_url:
            public_url = _best_company_photo_url(row, api_key=api_key_for_photo) or ""

        cache[key] = public_url
        return public_url
    except Exception:
        return ""


def _insert_excel_image(ws, cell_ref: str, img_bytes: bytes, mime_type: str = "image/jpeg", max_width_px: int = 150, max_height_px: int = 100):
    """Insere imagem no Excel ancorada na célula indicada.

    Observação importante: o openpyxl não lida bem com PIL.Image já redimensionada/copied,
    porque em algumas situações o objeto perde o atributo interno ``fp`` e o save do workbook
    quebra com: ``'Image' object has no attribute 'fp'``.
    Por isso, aqui a imagem é sempre normalizada e regravada em um BytesIO antes de virar XLImage.
    """
    if not img_bytes or PILImage is None:
        return False
    try:
        src_bio = BytesIO(img_bytes)
        pil = PILImage.open(src_bio)
        pil.load()

        width, height = pil.size
        if width <= 0 or height <= 0:
            return False

        scale = min(max_width_px / float(width), max_height_px / float(height), 1.0)
        new_w = max(1, int(width * scale))
        new_h = max(1, int(height * scale))

        if pil.mode not in ("RGB", "RGBA"):
            pil = pil.convert("RGBA" if "A" in getattr(pil, 'getbands', lambda: [])() else "RGB")

        if scale < 1.0:
            pil = pil.copy()
            pil.thumbnail((new_w, new_h), resample=getattr(PILImage, 'LANCZOS', 1))

        has_alpha = "A" in getattr(pil, 'getbands', lambda: [])()
        out_bio = BytesIO()
        if has_alpha:
            pil.save(out_bio, format="PNG")
            out_bio.name = "image.png"
        else:
            if pil.mode != "RGB":
                pil = pil.convert("RGB")
            pil.save(out_bio, format="JPEG", quality=92)
            out_bio.name = "image.jpg"
        out_bio.seek(0)

        img = XLImage(out_bio)
        img.width = new_w
        img.height = new_h
        img.anchor = cell_ref
        ws.add_image(img)

        col_letters = ''.join(ch for ch in cell_ref if ch.isalpha())
        row_digits = ''.join(ch for ch in cell_ref if ch.isdigit())
        if col_letters:
            current_w = ws.column_dimensions[col_letters].width
            desired_w = max((current_w or 8.43), round((new_w + 20) / 7, 2))
            ws.column_dimensions[col_letters].width = desired_w
        if row_digits:
            row_idx = int(row_digits)
            current_h = ws.row_dimensions[row_idx].height
            desired_h = max((current_h or 15), round((new_h * 0.78) + 8, 1))
            ws.row_dimensions[row_idx].height = desired_h
        return True
    except Exception:
        return False

def export_results_incremental_xlsx(
    df,
    template_bytes: bytes,
    sheet_name: str,
    executiva: str = "Vanessa",
    updated_dt: datetime = None,
    dedup: bool = True,
    google_api_key_for_photo: str = "",
):
    """
    Exporta incrementalmente para Excel no layout:
    Segmento | Cliente | Responsável | Contato | Já fiz contato? | Data de contato |
    E-mail | Endereço | CEP | Atualizado em | Executiva | Foto

    A foto é inserida como imagem embutida no .xlsx usando a mesma lógica visual da tela.
    """
    wb = load_workbook(BytesIO(template_bytes))
    if sheet_name not in wb.sheetnames:
        sheet_name = wb.sheetnames[0]
    ws = wb[sheet_name]

    header_row, col_map = _find_header_row(ws, scan_rows=40, scan_cols=80)

    # Migração de template antigo (com Bairro/Cidade) para o novo layout:
    # Segmento | Cliente | Responsável | Contato | Já fiz contato? | Data de contato | E-mail | Endereço | CEP | Atualizado em | Executiva | Foto
    try:
        old_headers = ["Segmento", "Cliente", "Responsável", "Contato", "E-mail", "Endereço", "Bairro", "Cidade", "CEP", "Atualizado em", "Executiva"]
        expected_headers = _EXPORT_REQUIRED_HEADERS

        if header_row is not None:
            cur = [ws.cell(header_row, c).value for c in range(1, 13)]
            cur_norm = [_norm_header(x) for x in cur]
            old_norm = [_norm_header(x) for x in old_headers]

            cur_trim = [x for x in cur_norm if x]
            if cur_trim == old_norm:
                ws.insert_cols(5, 2)   # após Contato
                ws.delete_cols(10, 1)  # remove Cidade (mantém Bairro)
                for i, h in enumerate(expected_headers, start=1):
                    ws.cell(header_row, i).value = h
                col_map = {}
                for c in range(1, 61):
                    v = _norm_header(ws.cell(header_row, c).value)
                    if v:
                        col_map.setdefault(v, []).append(c)
    except Exception:
        pass

    if header_row is None:
        header_row = 1
        for i, h in enumerate(_EXPORT_REQUIRED_HEADERS, start=1):
            ws.cell(header_row, i).value = h
        col_map = {_norm_header(h): [i] for i, h in enumerate(_EXPORT_REQUIRED_HEADERS, start=1)}
    else:
        all_cols = col_map

        def pick(h):
            cols = all_cols.get(_norm_header(h)) or []
            return cols[0] if cols else None

        def first_empty_col(max_col=30):
            for c in range(1, max_col + 1):
                v = ws.cell(header_row, c).value
                if v is None or str(v).strip() == "":
                    return c
            return (ws.max_column or max_col) + 1

        for h in _EXPORT_REQUIRED_HEADERS:
            if pick(h) is None:
                c = first_empty_col()
                ws.cell(header_row, c).value = h
                all_cols.setdefault(_norm_header(h), []).append(c)
        col_map = all_cols

    def col(h):
        cols = col_map.get(_norm_header(h)) or []
        return cols[0] if cols else None

    cols = {h: col(h) for h in _EXPORT_REQUIRED_HEADERS}

    if updated_dt is None:
        if ZoneInfo is not None:
            updated_dt = datetime.now(ZoneInfo("America/Sao_Paulo"))
        else:
            updated_dt = datetime.now()
    updated_str = updated_dt.strftime("%d/%m/%Y %H:%M:%S")
    google_api_key_for_photo = (google_api_key_for_photo or os.getenv("GOOGLE_API_KEY", "") or _safe_secrets_get("GOOGLE_API_KEY", "") or "").strip()

    last_existing = _last_filled_row_any(ws, header_row, [cols["Cliente"]])

    existing = set()
    if dedup:
        for r in range(header_row + 1, last_existing + 1):
            c = ws.cell(r, cols["Cliente"]).value
            t = ws.cell(r, cols["Contato"]).value
            key = _dedup_key(str(c or ""), str(t or ""), str(ws.cell(r, cols["E-mail"]).value if cols.get("E-mail") else ""), str(ws.cell(r, cols["Endereço"]).value if cols.get("Endereço") else ""))
            if key != ("", ""):
                existing.add(key)

    start_row = max(header_row + 1, last_existing + 1)

    def extract_email(*vals) -> str:
        for v in vals:
            if v is None:
                continue
            s = str(v).strip()
            mm = re.search(r"([A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,})", s)
            if mm:
                return mm.group(1)
        return ""

    added = 0
    skipped = 0
    pending_photos = {}

    for _, row in df.iterrows():
        segmento = str(_pick_row_value(row, ["Categoria", "Segmento"])).strip()
        cliente = str(_pick_row_value(row, ["Nome", "Cliente", "Estabelecimento", "Nome do estabelecimento"])).strip()

        # Contato robusto (telefones + sociais + site) e e-mail quando existir
        email = _get_email_or_site(row)
        contato = _get_primary_phone_ddd(row)
        endereco, bairro, _cidade, cep = _get_endereco_fields(row)
        responsavel = _get_responsavel_row(row)

        if contato == "-" or contato.lower() == "nan":
            contato = ""
        if cliente.lower() == "nan":
            cliente = ""

        key = _dedup_key(cliente, contato, email, endereco)
        if dedup and key in existing and key != ("", ""):
            skipped += 1
            continue

        foto_bytes = None
        if google_api_key_for_photo:
            foto_bytes, _foto_mime = _get_photo_bytes_for_export(row, google_api_key_for_photo)
            if foto_bytes and key != ("", ""):
                pending_photos[key] = foto_bytes

        rr = start_row + added
        ws.cell(rr, cols["Segmento"]).value = segmento
        ws.cell(rr, cols["Cliente"]).value = cliente
        ws.cell(rr, cols["Responsável"]).value = responsavel
        ws.cell(rr, cols["Contato"]).value = contato
        ws.cell(rr, cols["Já fiz contato?"]).value = ""
        ws.cell(rr, cols["Data de contato"]).value = ""
        ws.cell(rr, cols["E-mail"]).value = email
        ws.cell(rr, cols["Endereço"]).value = endereco
        ws.cell(rr, cols["CEP"]).value = cep
        ws.cell(rr, cols["Atualizado em"]).value = updated_str
        ws.cell(rr, cols["Executiva"]).value = executiva
        ws.cell(rr, cols["Foto"]).value = "Imagem" if (foto_bytes if "foto_bytes" in locals() else None) else ""

        added += 1
        if dedup and key != ("", ""):
            existing.add(key)

    # Ordenar por Segmento somente quando a aba ainda não tiver imagens, para não desalinhar fotos já embutidas.
    try:
        existing_images = list(getattr(ws, "_images", []) or [])
        if added > 0 and not pending_photos and not existing_images:
            seg_col = cols.get("Segmento")
            if seg_col:
                last_row = _last_filled_row_any(ws, header_row, [cols.get("Cliente"), cols.get("Contato"), seg_col])
                data_rows = []
                for r in range(header_row + 1, last_row + 1):
                    row_vals = [ws.cell(r, cols[h]).value if cols.get(h) else None for h in _EXPORT_REQUIRED_HEADERS]
                    data_rows.append(row_vals)
                data_rows.sort(key=lambda rv: str(rv[0] or "").strip().lower())
                for i, rv in enumerate(data_rows, start=header_row + 1):
                    for j, h in enumerate(_EXPORT_REQUIRED_HEADERS):
                        c = cols.get(h)
                        if c:
                            ws.cell(i, c).value = rv[j]
    except Exception:
        pass
    # Formatação: Data de contato (dd/mm/aaaa) e validação simples para "Já fiz contato?"
    try:
        date_col = cols.get("Data de contato")
        chk_col = cols.get("Já fiz contato?")
        if date_col:
            for r in range(header_row + 1, (ws.max_row or header_row) + 1):
                ws.cell(r, date_col).number_format = "dd/mm/yyyy"
        if chk_col:
            from openpyxl.worksheet.datavalidation import DataValidation
            from openpyxl.utils import get_column_letter
            dv = DataValidation(type="list", formula1='"TRUE,FALSE"', allow_blank=True)
            ws.add_data_validation(dv)
            dv.add(f"{get_column_letter(chk_col)}{header_row+1}:{get_column_letter(chk_col)}{ws.max_row or header_row+1}")
    except Exception:
        pass


    # Inserção de imagens no Excel após a escrita das linhas.
    try:
        foto_col = cols.get("Foto")
        if foto_col and pending_photos:
            for r in range(header_row + 1, (ws.max_row or header_row) + 1):
                cliente0 = ws.cell(r, cols["Cliente"]).value if cols.get("Cliente") else ""
                contato0 = ws.cell(r, cols["Contato"]).value if cols.get("Contato") else ""
                email0 = ws.cell(r, cols["E-mail"]).value if cols.get("E-mail") else ""
                endereco0 = ws.cell(r, cols["Endereço"]).value if cols.get("Endereço") else ""
                row_key = _dedup_key(str(cliente0 or ""), str(contato0 or ""), str(email0 or ""), str(endereco0 or ""))
                img_bytes = pending_photos.get(row_key)
                if not img_bytes:
                    continue
                cell_ref = f"{ws.cell(header_row, foto_col).column_letter}{r}"
                ok = _insert_excel_image(ws, cell_ref, img_bytes, "image/jpeg")
                if ok:
                    ws.cell(r, foto_col).value = "Imagem"
    except Exception:
        pass

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return out.getvalue(), {"added": added, "skipped": skipped, "sheet": sheet_name}

def _parse_gsheet_id(url_or_id: str) -> str:
    s = (url_or_id or "").strip()
    if not s:
        return ""
    if re.fullmatch(r"[a-zA-Z0-9-_]{20,}", s):
        return s
    m = re.search(r"/spreadsheets/d/([a-zA-Z0-9-_]+)", s)
    if m:
        return m.group(1)
    return ""


def _extract_sheet_id(url_or_id: str) -> str:
    """
    Aceita URL completa do Google Sheets ou ID puro e retorna somente o spreadsheet_id.
    Alias para compatibilidade com versões anteriores.
    """
    s = str(url_or_id or "").strip()
    if not s:
        return ""
    try:
        return _parse_gsheet_id(s)
    except Exception:
        pass
    m = re.search(r"/spreadsheets/d/([a-zA-Z0-9-_]+)", s)
    if m:
        return m.group(1)
    return s


def _get_gspread_client(sa_info_override: dict | None = None, cred_path_override: str | None = None):
    """
    Streamlit Cloud:
      - coloque o JSON do Service Account em st.secrets, ex:
        [gcp_service_account]
        type="service_account"
        project_id="..."
        private_key="-----BEGIN PRIVATE KEY-----\n...\n-----END PRIVATE KEY-----\n"
        client_email="..."
        ...
    Alternativas:
      - env var GOOGLE_SERVICE_ACCOUNT_JSON (json string)
      - env var GOOGLE_APPLICATION_CREDENTIALS (caminho do json)
    """
    if gspread is None or Credentials is None:
        raise RuntimeError("Dependências não instaladas. Instale: gspread e google-auth (ou adicione no requirements.txt).")

    # Override (uso local): JSON enviado via UI ou caminho digitado pelo usuário
    if sa_info_override:
        scopes = ["https://www.googleapis.com/auth/spreadsheets", "https://www.googleapis.com/auth/drive.file"]
        creds = Credentials.from_service_account_info(sa_info_override, scopes=scopes)
        return gspread.authorize(creds)


    sa_info = None
    try:
        if _safe_secrets_has("gcp_service_account"):
            sa_info = dict(_safe_secrets_get("gcp_service_account"))
        elif _safe_secrets_has("google_service_account"):
            sa_info = dict(_safe_secrets_get("google_service_account") or {})
        elif _safe_secrets_has("service_account"):
            sa_info = dict(_safe_secrets_get("service_account") or {})
    except Exception:
        sa_info = None

    if sa_info:
        scopes = ["https://www.googleapis.com/auth/spreadsheets", "https://www.googleapis.com/auth/drive.file"]
        creds = Credentials.from_service_account_info(sa_info, scopes=scopes)
        return gspread.authorize(creds)

    env_json = os.getenv("GOOGLE_SERVICE_ACCOUNT_JSON", "").strip()
    if env_json:
        import json as _json
        sa_info = _json.loads(env_json)
        scopes = ["https://www.googleapis.com/auth/spreadsheets", "https://www.googleapis.com/auth/drive.file"]
        creds = Credentials.from_service_account_info(sa_info, scopes=scopes)
        return gspread.authorize(creds)

    cred_path = (cred_path_override or os.getenv("GOOGLE_APPLICATION_CREDENTIALS", "") or "").strip()
    if cred_path and os.path.exists(cred_path):
        return gspread.service_account(filename=cred_path)

    # Uso local: se existir um JSON na mesma pasta do app, usa automaticamente
    try:
        local_fp = os.path.join(_app_base_dir(), "service_account.json")
        if os.path.exists(local_fp):
            return gspread.service_account(filename=local_fp)
    except Exception:
        pass

    raise RuntimeError(
        "Credenciais do Google não configuradas. "
        "Opções: (1) Cloud: colocar o Service Account JSON em st.secrets (gcp_service_account); "
        "(2) Local: enviar o JSON na tela do app; "
        "(3) Local: definir GOOGLE_APPLICATION_CREDENTIALS apontando para o .json; "
        "(4) Local: colocar service_account.json na mesma pasta do app."
    )


def _gs_norm(v) -> str:
    return ("" if v is None else str(v)).strip().lower()


# ===============================
# GOOGLE DRIVE (para salvar fotos e usar =IMAGE no Sheets)
# ===============================

def _get_drive_service(sa_info_override: dict | None = None, cred_path_override: str | None = None):
    """Cria um service do Google Drive usando o mesmo Service Account do Sheets."""
    scopes = ["https://www.googleapis.com/auth/spreadsheets", "https://www.googleapis.com/auth/drive.file"]

    if sa_info_override:
        creds = Credentials.from_service_account_info(sa_info_override, scopes=scopes)
        return build("drive", "v3", credentials=creds, cache_discovery=False)

    sa_info = None
    try:
        if _safe_secrets_has("gcp_service_account"):
            sa_info = dict(_safe_secrets_get("gcp_service_account"))
        elif _safe_secrets_has("google_service_account"):
            sa_info = dict(_safe_secrets_get("google_service_account"))
        elif _safe_secrets_has("service_account"):
            sa_info = dict(_safe_secrets_get("service_account") or {})
    except Exception:
        sa_info = None

    if sa_info:
        creds = Credentials.from_service_account_info(sa_info, scopes=scopes)
        return build("drive", "v3", credentials=creds, cache_discovery=False)

    env_json = os.getenv("GOOGLE_SERVICE_ACCOUNT_JSON", "").strip()
    if env_json:
        import json as _json
        sa_info = _json.loads(env_json)
        creds = Credentials.from_service_account_info(sa_info, scopes=scopes)
        return build("drive", "v3", credentials=creds, cache_discovery=False)

    if cred_path_override and os.path.exists(cred_path_override):
        creds = Credentials.from_service_account_file(cred_path_override, scopes=scopes)
        return build("drive", "v3", credentials=creds, cache_discovery=False)

    env_path = os.getenv("GOOGLE_APPLICATION_CREDENTIALS", "").strip()
    if env_path and os.path.exists(env_path):
        creds = Credentials.from_service_account_file(env_path, scopes=scopes)
        return build("drive", "v3", credentials=creds, cache_discovery=False)

    try:
        local_fp = os.path.join(_app_base_dir(), "service_account.json")
        if os.path.exists(local_fp):
            creds = Credentials.from_service_account_file(local_fp, scopes=scopes)
            return build("drive", "v3", credentials=creds, cache_discovery=False)
    except Exception:
        pass

    for p in ["service_account.json", "credentials.json", "sa.json"]:
        if os.path.exists(p):
            creds = Credentials.from_service_account_file(p, scopes=scopes)
            return build("drive", "v3", credentials=creds, cache_discovery=False)

    raise RuntimeError("Service Account não encontrado para Drive (st.secrets ou arquivo local).")


def _drive_get_or_create_folder_id(
    folder_name: str,
    sa_info_override: dict | None = None,
    cred_path_override: str | None = None,
) -> str:
    """Cria/obtém uma pasta no Drive do Service Account."""
    try:
        if "drive_folder_cache" not in st.session_state:
            st.session_state.drive_folder_cache = {}
        cache: dict = st.session_state.drive_folder_cache

        cred_key = "default"
        if sa_info_override:
            cred_key = str((sa_info_override.get("client_email") or "override")).strip().lower()
        elif cred_path_override:
            cred_key = os.path.abspath(cred_path_override)

        cache_key = f"{cred_key}|{folder_name.strip().lower()}"
        if cache_key in cache and cache[cache_key]:
            return cache[cache_key]
    except Exception:
        cache = {}
        cache_key = folder_name.strip().lower()

    svc = _get_drive_service(sa_info_override=sa_info_override, cred_path_override=cred_path_override)
    safe_name = folder_name.replace("'", "\'")
    q = (
        f"name='{safe_name}' and "
        "mimeType='application/vnd.google-apps.folder' and trashed=false"
    )
    res = svc.files().list(q=q, fields="files(id,name)", pageSize=10).execute()
    files = res.get("files", [])
    if files:
        folder_id = files[0]["id"]
        try:
            cache[cache_key] = folder_id
        except Exception:
            pass
        return folder_id

    meta = {"name": folder_name, "mimeType": "application/vnd.google-apps.folder"}
    created = svc.files().create(body=meta, fields="id").execute()
    folder_id = created["id"]
    try:
        cache[cache_key] = folder_id
    except Exception:
        pass
    return folder_id


def _drive_upload_image_and_get_public_url(
    image_bytes: bytes,
    filename: str,
    folder_id: str,
    mime_type: str = "image/jpeg",
    sa_info_override: dict | None = None,
    cred_path_override: str | None = None,
) -> str:
    """Upload da imagem no Drive e retorna um link direto para =IMAGE no Sheets."""
    svc = _get_drive_service(sa_info_override=sa_info_override, cred_path_override=cred_path_override)

    media = MediaIoBaseUpload(io.BytesIO(image_bytes), mimetype=mime_type, resumable=False)
    file_metadata = {"name": filename, "parents": [folder_id]}
    created = svc.files().create(body=file_metadata, media_body=media, fields="id").execute()
    file_id = created["id"]

    try:
        svc.permissions().create(
            fileId=file_id,
            body={"type": "anyone", "role": "reader"},
            fields="id",
        ).execute()
    except Exception:
        pass

    return f"https://drive.google.com/uc?export=view&id={file_id}"


def _download_image_bytes(url: str) -> tuple[bytes, str] | tuple[None, None]:
    """Baixa a imagem e retorna (bytes, mime_type)."""
    try:
        r = requests.get(url, timeout=25, allow_redirects=True, stream=True)
        if r.status_code != 200:
            return (None, None)
        ctype = (r.headers.get("Content-Type") or "").split(";")[0].strip().lower()
        if not ctype.startswith("image/"):
            return (None, None)
        content = r.content
        if not content:
            return (None, None)
        return (content, ctype or "image/jpeg")
    except Exception:
        return (None, None)


def _photo_cell_formula_for_row(
    row,
    api_key_for_photo: str,
    sa_info_override: dict | None = None,
    cred_path_override: str | None = None,
) -> str:
    """Retorna a fórmula =IMAGE(...) para o Google Sheets.

    Mantém a estratégia mais estável:
    1) tenta o fluxo já validado do Drive para o Sheets;
    2) se falhar, usa a URL pública direta da foto.
    """
    try:
        api_key_for_photo = (api_key_for_photo or "").strip()
        if not api_key_for_photo:
            return ""

        if "photo_cache" not in st.session_state:
            st.session_state.photo_cache = {}
        cache: dict = st.session_state.photo_cache

        place_id = ""
        if hasattr(row, "get"):
            place_id = str(row.get("PlaceId") or row.get("place_id") or "").strip()
        nome = str(row.get("Nome") if hasattr(row, "get") else "")[:80].strip() or "empresa"
        endereco_key = str(row.get("Endereço") if hasattr(row, "get") else "")
        key = place_id or hashlib.md5((nome + endereco_key).encode("utf-8", errors="ignore")).hexdigest()

        if key in cache:
            url = cache[key]
            return f'=IMAGE("{url}")' if url else ""

        img_bytes, mime = _get_photo_bytes_for_export(row, api_key_for_photo)
        if img_bytes:
            folder_id = _drive_get_or_create_folder_id(
                "Fotos - Info Comercial",
                sa_info_override=sa_info_override,
                cred_path_override=cred_path_override,
            )
            ext = "jpg"
            if mime == "image/png":
                ext = "png"
            elif mime == "image/webp":
                ext = "webp"
            filename = f"{nome} - {key}.{ext}"

            public_url = _drive_upload_image_and_get_public_url(
                img_bytes,
                filename,
                folder_id,
                mime_type=mime,
                sa_info_override=sa_info_override,
                cred_path_override=cred_path_override,
            )
            if public_url:
                cache[key] = public_url
                return f'=IMAGE("{public_url}")'

        fallback_url = _photo_public_url_for_row(row, api_key_for_photo)
        cache[key] = fallback_url or ""
        return f'=IMAGE("{fallback_url}")' if fallback_url else ""
    except Exception:
        return ""


def _digits_only(s: str) -> str:
    return "".join(ch for ch in (s or "") if ch.isdigit())



def _gs_sort_by_segmento(ws, data_rows_count: int):
    """Ordena a aba pela melhor coluna de segmento disponível, mantendo o header na linha 1."""
    try:
        if not data_rows_count or int(data_rows_count) <= 1:
            return

        header = ws.row_values(1) or []
        header_norm = [_gs_norm(x) for x in header]

        sort_idx = 0
        if "segmento_txt" in header_norm:
            sort_idx = header_norm.index("segmento_txt")
        elif "segmento" in header_norm:
            sort_idx = header_norm.index("segmento")

        end_col = max(len(header), len(_GS_EXPORT_HEADERS))

        ws.spreadsheet.batch_update(
            {
                "requests": [
                    {
                        "sortRange": {
                            "range": {
                                "sheetId": ws.id,
                                "startRowIndex": 1,
                                "endRowIndex": 1 + int(data_rows_count),
                                "startColumnIndex": 0,
                                "endColumnIndex": end_col,
                            },
                            "sortSpecs": [{"dimensionIndex": int(sort_idx), "sortOrder": "ASCENDING"}],
                        }
                    }
                ]
            }
        )
    except Exception:
        return



def _gs_apply_contact_columns(ws, max_rows: int = 5000):
    """
    Aplica:
    - Checkbox na coluna "Já fiz contato?"
    - Formato de data dd/MM/yyyy na coluna "Data de contato"
    """
    try:
        header = ws.row_values(1) or []
        header_norm = [_gs_norm(x) for x in header]

        def idx_of(name: str, default_1based: int):
            n = _gs_norm(name)
            if n in header_norm:
                return header_norm.index(n) + 1
            return default_1based

        col_checkbox = idx_of("Já fiz contato?", 9)
        col_date = idx_of("Data de contato", 10)

        end_row = int(max_rows)

        cb_start_col = col_checkbox - 1
        cb_end_col = col_checkbox
        dt_start_col = col_date - 1
        dt_end_col = col_date

        ws.spreadsheet.batch_update({
            "requests": [
                {
                    "setDataValidation": {
                        "range": {
                            "sheetId": ws.id,
                            "startRowIndex": 1,
                            "endRowIndex": end_row,
                            "startColumnIndex": cb_start_col,
                            "endColumnIndex": cb_end_col
                        },
                        "rule": {
                            "condition": {"type": "BOOLEAN"},
                            "showCustomUi": True,
                            "strict": True
                        }
                    }
                },
                {
                    "repeatCell": {
                        "range": {
                            "sheetId": ws.id,
                            "startRowIndex": 1,
                            "endRowIndex": end_row,
                            "startColumnIndex": dt_start_col,
                            "endColumnIndex": dt_end_col
                        },
                        "cell": {
                            "userEnteredFormat": {
                                "numberFormat": {"type": "DATE", "pattern": "dd/MM/yyyy"}
                            }
                        },
                        "fields": "userEnteredFormat.numberFormat"
                    }
                }
            ]
        })
    except Exception:
        return

def _ensure_gsheet_headers(ws):
    """Garante apenas o layout de C:R, sem alterar nada nas colunas A e B."""
    expected_dynamic = _GS_DYNAMIC_HEADERS
    ws.update("C1:R1", [expected_dynamic])

def export_results_incremental_gsheet(
    df,
    spreadsheet_url_or_id: str,
    worksheet_name: str = DEFAULT_GSHEET_TAB,
    executiva: str = "Vanessa",
    updated_dt: datetime = None,
    dedup: bool = True,
    sa_info_override: dict | None = None,
    cred_path_override: str | None = None,
    google_api_key_for_photo: str = "",
):
    """
    Exporta incrementalmente para Google Sheets no layout:
    ID_CLIENTE | Segmento | Segmento_TXT | Executiva Pixel | Empresa (Cliente) |
    Responsável pela Empresa | Telefone | E-mail | Já fiz contato? | Data de contato |
    Observações | Site | Endereço | Bairro | CEP | Atualizado em | Foto | Foto_AppSheet

    As colunas A/B não são alteradas.
    A exportação escreve somente de C a R.
    A coluna R recebe apenas a URL direta da imagem (Foto_AppSheet).
    """
    gc = _get_gspread_client(sa_info_override=sa_info_override, cred_path_override=cred_path_override)

    spreadsheet_id = _extract_sheet_id(spreadsheet_url_or_id or DEFAULT_GSHEET_ID)
    if not spreadsheet_id:
        raise RuntimeError("ID da planilha não informado (GSHEET_ID/DEFAULT_GSHEET_ID).")

    sh = gc.open_by_key(spreadsheet_id)
    ws = sh.worksheet(worksheet_name)

    _ensure_gsheet_headers(ws)
    _gs_apply_contact_columns(ws)

    if updated_dt is None:
        if ZoneInfo is not None:
            updated_dt = datetime.now(ZoneInfo("America/Sao_Paulo"))
        else:
            updated_dt = datetime.now()
    updated_str = updated_dt.strftime("%d/%m/%Y %H:%M:%S")

    google_api_key_for_photo = (google_api_key_for_photo or os.getenv("GOOGLE_API_KEY", "") or _safe_secrets_get("GOOGLE_API_KEY", "") or "").strip()

    existing_rows = ws.get("A2:R") or []

    existing = set()
    existing_row_map = {}
    if dedup:
        for idx, r in enumerate(existing_rows, start=2):
            cliente0 = (r[4] if len(r) > 4 else "").strip()
            contato0 = (r[6] if len(r) > 6 else "").strip()
            email0 = (r[7] if len(r) > 7 else "").strip()
            endereco0 = (r[12] if len(r) > 12 else "").strip()
            foto0 = (r[16] if len(r) > 16 else "").strip()
            foto_app0 = (r[17] if len(r) > 17 else "").strip()
            key = _dedup_key(cliente0, contato0, email0, endereco0)
            if key != ("", ""):
                existing.add(key)
                if key not in existing_row_map:
                    existing_row_map[key] = {
                        "row_number": idx,
                        "foto": foto0,
                        "foto_appsheet": foto_app0,
                    }

    rows_to_append = []
    skipped = 0
    cells_to_update = []

    for _, row in df.iterrows():
        segmento_txt = str(_pick_row_value(row, ["Categoria", "Segmento", "Segmento_TXT"])).strip()
        cliente = str(_pick_row_value(row, ["Nome", "Cliente", "Estabelecimento", "Nome do estabelecimento", "Empresa (Cliente)"])).strip()
        telefone = _get_primary_phone_ddd(row)
        email = _get_email_row(row)
        site = _get_site_row(row)
        endereco, bairro, _cidade, cep = _get_endereco_fields(row)
        foto_cell = _photo_cell_formula_for_row(
            row,
            api_key_for_photo=google_api_key_for_photo,
            sa_info_override=sa_info_override,
            cred_path_override=cred_path_override,
        )
        foto_appsheet = _photo_public_url_for_row(
            row,
            api_key_for_photo=google_api_key_for_photo,
        )
        responsavel = _get_responsavel_row(row)

        key = _dedup_key(cliente, telefone, email, endereco)
        if dedup and key in existing and key != ("", ""):
            existing_info = existing_row_map.get(key) or {}
            row_number = existing_info.get("row_number")
            foto_atual = (existing_info.get("foto") or "").strip()
            foto_app_atual = (existing_info.get("foto_appsheet") or "").strip()

            if row_number:
                if foto_cell and not foto_atual:
                    cells_to_update.append({"range": f"Q{row_number}", "values": [[foto_cell]]})
                    existing_info["foto"] = foto_cell
                if foto_appsheet and not foto_app_atual:
                    cells_to_update.append({"range": f"R{row_number}", "values": [[foto_appsheet]]})
                    existing_info["foto_appsheet"] = foto_appsheet

            skipped += 1
            continue

        rows_to_append.append([
            segmento_txt,
            executiva,
            cliente,
            responsavel,
            telefone,
            email,
            False,
            "",
            "",
            site,
            endereco,
            bairro,
            cep,
            updated_str,
            foto_cell,
            foto_appsheet,
        ])
        if dedup and key != ("", ""):
            existing.add(key)

    if rows_to_append:
        ws.append_rows(rows_to_append, value_input_option="USER_ENTERED", table_range="C:R")
        _gs_sort_by_segmento(ws, data_rows_count=(len(existing_rows) + len(rows_to_append)))

    if cells_to_update:
        ws.batch_update(cells_to_update, value_input_option="USER_ENTERED")

    return {
        "added": len(rows_to_append),
        "updated_photo_cells": len(cells_to_update),
        "skipped": skipped,
        "spreadsheet_id": spreadsheet_id,
        "worksheet": worksheet_name,
    }

def nominatim_suggest(query: str, limit: int = 6) -> list[dict]:
    """
    Busca sugestões no Nominatim (OSM).
    Retorna uma lista de dicts com display_name + address details.
    """
    q = (query or "").strip()
    if len(q) < 4:
        return []
    url = "https://nominatim.openstreetmap.org/search"
    params = {
        "format": "jsonv2",
        "q": q,
        "addressdetails": 1,
        "limit": int(limit),
        "countrycodes": "br",
    }
    headers = {"User-Agent": "SuplemexApp/1.0 (streamlit)"}
    try:
        r = requests.get(url, params=params, headers=headers, timeout=15)
        r.raise_for_status()
        data = r.json()
        if not isinstance(data, list):
            return []
        return data
    except Exception:
        return []

def _uf_allowed_for_result(res: dict) -> bool:
    if not ALLOWED_UF:
        return True
    try:
        addr = res.get("address") or {}
        state = (addr.get("state") or "").strip()
        # aceita se o estado bater com algum permitido
        allowed_states = set((v or "").strip().lower() for v in ALLOWED_UF.values())
        return (state.lower() in allowed_states) or any(a in state.lower() for a in allowed_states)
    except Exception:
        return False

def google_autocomplete_suggest(query: str, api_key: str, limit: int = 6) -> list[str]:
    """
    Busca sugestões de endereço via Google Places Autocomplete.
    Usa a mesma família de APIs já utilizada no app para Nearby/Details.
    """
    q = (query or "").strip()
    if len(q) < 4 or not api_key:
        return []
    url = "https://maps.googleapis.com/maps/api/place/autocomplete/json"
    params = {
        "input": q,
        "language": "pt-BR",
        "components": "country:br",
        "types": "address",
        "key": api_key,
    }
    try:
        r = requests.get(url, params=params, timeout=15)
        r.raise_for_status()
        data = r.json() or {}
        preds = data.get("predictions") or []
        out = []
        for item in preds[: max(1, int(limit))]:
            desc = str(item.get("description") or "").strip()
            if desc:
                out.append(desc)
        return out
    except Exception:
        return []


def get_suggestions_filtered(query: str, limit: int = 6, api_key: str = "") -> list[str]:
    out: list[str] = []

    # 1) Tenta Google Autocomplete primeiro (mais estável no Cloud e coerente com o restante do app)
    google_sugs = google_autocomplete_suggest(query, api_key=api_key, limit=limit)
    out.extend(google_sugs)

    # 2) Fallback Nominatim/OSM
    if len(out) < limit:
        raw = nominatim_suggest(query, limit=limit)
        for r in raw:
            if not isinstance(r, dict):
                continue
            if not _uf_allowed_for_result(r):
                continue
            dn = (r.get("display_name") or "").strip()
            if dn:
                out.append(dn)

    # remove duplicados preservando ordem
    seen = set()
    uniq = []
    for s in out:
        k = s.lower()
        if k in seen:
            continue
        seen.add(k)
        uniq.append(s)
    return uniq[: max(1, int(limit))]

def _highlight_match(texto: str, termo: str) -> str:
    """Destaca o termo digitado em negrito (HTML), mantendo segurança."""
    t = (texto or "")
    q = (termo or "").strip()
    if not q:
        return html.escape(t)
    esc = html.escape(t)
    # destaca todas as ocorrências (case-insensitive)
    try:
        pattern = re.compile(re.escape(q), re.IGNORECASE)
        return pattern.sub(lambda m: f"<b>{html.escape(m.group(0))}</b>", esc)
    except Exception:
        return esc


def _split_suggestion(s: str) -> tuple[str, str]:
    """Divide sugestão longa em linha principal + complementar (para UI mais limpa)."""
    s = (s or "").strip()
    if not s:
        return "", ""
    parts = [p.strip() for p in s.split(",") if p.strip()]
    if len(parts) <= 3:
        return s, ""
    main = ", ".join(parts[:3])
    sub = ", ".join(parts[3:])
    return main, sub



def main():

    # ===============================
    # UI / STYLE
    # ===============================
    st.set_page_config(page_title="ProSearch", page_icon="🗺️", layout="wide", initial_sidebar_state="expanded")

    st.markdown("""
        <style>
        /* Força tema dark */
        :root{
        --bg:#0b1220;
        --card:#0f172a;
        --stroke:#1f2a44;
        --text:#e5e7eb;
        --muted:#a1a1aa;
        --brand:#60a5fa;
        }
        .stApp{ background: var(--bg) !important; color: var(--text) !important; }

        /* Sidebar */
        section[data-testid="stSidebar"]{
        background: #0a1020 !important;
        border-right: 1px solid var(--stroke) !important;
        }

        /* Containers e elementos do Streamlit */
        div[data-testid="stVerticalBlock"]{ color: var(--text) !important; }
        div[data-testid="stMarkdownContainer"]{ color: var(--text) !important; }
        div[data-testid="stMetricValue"]{ color: var(--text) !important; }

        /* Inputs */
        div[data-baseweb="input"] input,
        div[data-baseweb="textarea"] textarea{
        background: #0b1220 !important;
        color: var(--text) !important;
        border: 1px solid var(--stroke) !important;
        }
        div[data-baseweb="select"] > div{
        background: #0b1220 !important;
        color: var(--text) !important;
        border: 1px solid var(--stroke) !important;
        }

        /* Botões */
        button[kind="primary"]{
        background: #2563eb !important;
        color: white !important;
        border: 1px solid #2563eb !important;
        }
        button{
        border-radius: 12px !important;
        border: 1px solid var(--stroke) !important;
        }

        /* Dataframe (melhora contraste) */
        div[data-testid="stDataFrame"]{
        border: 1px solid var(--stroke) !important;
        border-radius: 12px !important;
        overflow: hidden !important;
        }
        </style>
        """, unsafe_allow_html=True)


    # ===============================
    # CATEGORIAS (Google type + OSM tags)
    # ===============================
    CATEGORIES_CONFIG = {
        "🛒 Mercados": {"google_type": "supermarket", "osm": ["shop=supermarket", "shop=convenience", "shop=grocery"]},
        "🏫 Escolas": {"google_type": ["school", "primary_school", "secondary_school"], "osm": ["amenity=school", "amenity=kindergarten", "amenity=university", "amenity=language_school"], "google_keywords": ["Escola", "Colégio", "Curso de inglês", "Escola de idiomas", "Maple Bear", "Maple Bear Canadian School", "CCAA", "Fisk", "CNA", "Wizard", "Wise Up", "KNN Idiomas", "Yázigi", "Yes! Idiomas", "Cultura Inglesa"], "name_exclude": ["estadual", "municipal", "ciep", "CIEP"]},
        "🏫 Faculdades/Universidades": {"google_type": "school", "osm": ["amenity=university"]},
        "✈️ Agências de Viagens": {
            "google_type": ["travel_agency", "tourist_information"],
            "osm": ["shop=travel_agency", "office=travel_agent", "tourism=information"],
            "google_keywords": [
                "Agência de Viagens", "Agencia de Viagens", "Turismo", "Viagens",
                "Pacotes", "Pacotes de viagem", "Passagens", "Passagens aéreas",
                "Passagens aereas", "Cruzeiro", "Cruzeiros", "Intercâmbio", "Intercambio",
                "Operadora de turismo", "Operadora", "Excursão", "Excursao"
            ],
            "segment_name_any": [
                "agência de viagens", "agencia de viagens", "turismo", "viagens",
                "passagens", "passagem", "pacote", "pacotes", "cruzeiro", "intercâmbio", "intercambio",
                "operadora", "excursão", "excursao"
            ],
            "name_exclude": ["hotel", "pousada", "hostel", "motel", "hospedagem"]
        },
        "🏥 Hospitais": {"google_type": "hospital", "osm": ["amenity=hospital", "amenity=clinic", "amenity=doctors"]},
        "💊 Farmácias": {"google_type": "pharmacy", "osm": ["amenity=pharmacy"]},
        "🚗 Automotivos": {"google_type": "car_dealer", "osm": ["amenity=car_dealer", "shop=car", "amenity=car_repair", "shop=car_repair"]},
        "🍽️ Restaurantes": {"google_type": "restaurant", "osm": ["amenity=restaurant", "amenity=cafe", "amenity=fast_food"]},
        "💪 Academias": {"google_type": "gym", "osm": ["leisure=fitness_centre", "leisure=sports_centre"]},
        "🚘 Concessionárias": {
            "google_type": "car_dealer",
            "osm": ["amenity=car_dealer", "shop=car"],
            "google_keywords": ["Concessionária Jeep", "Concessionária Fiat", "Concessionária Ford", "Concessionária Volkswagen", "Concessionária VW", "Concessionária Chevrolet", "Concessionária Hyundai", "Concessionária Nissan", "Concessionária Renault", "Concessionária Honda", "Concessionária Toyota","Concessionária Mitsubishi", "Concessionária Kia", "Concessionária Peugeot", "Concessionária Citroën", "Concessionária Suzuki", "Concessionária Jac", "Concessionária Byd", "Concessionária Chery", "Concessionária Lifan"],
            "name_must_contain": ["jeep", "fiat", "ford", "volkswagen", "vw", "chevrolet", "hyundai", "nissan", "renault", "honda", "toyota","mitsubishi", "kia", "peugeot", "citroën", "suzuki", "jac", "byd", "chery", "lifan"],
            "name_exclude": ["hotel", "pousada", "hostel", "hospedagem", "motel"]
        },
        "🏗️ Construtoras / Incorporadoras": {
            "google_type": ["general_contractor", "real_estate_agency"],
            "osm": ["office=construction", "craft=builder", "office=architect"],
            "google_keywords": ["Stand de vendas", "Estande de vendas", "Plantão de vendas", "Plantao de vendas", "Stand imobiliário", "Estande imobiliário", "Estande imobiliario", "Stand", "Estande", "Gafisa", "Mozak", "Mozak Rio", "LatinExclusive", "Latin Exclusive", "Incorporadora Gafisa", "Incorporadora Mozak", "Incorporadora Latin Exclusive", "Construtora MRV", "Construtora Direcional", "Construtora Tenda", "Construtora Cury", "Construtora Cyrela", "Construtora Even", "Construtora Gafisa","Construtora Rossi", "Construtora Trisul", "Construtora Eztec", "Construtora Tecnisa", "Construtora Brookfield", "Construtora Plaenge", "Construtora Mitre", "Construtora Viver", "Construtora Rodobens", "Construtora Patrimar", "Incorporadora", "Incorporação imobiliária", "Incorporadora MRV", "Incorporadora Direcional", "Incorporadora Tenda", "Incorporadora Cury", "Incorporadora Cyrela", "Incorporadora Even", "Incorporadora Gafisa", "Incorporadora Mitre", "Incorporadora Eztec"],
            "name_must_contain": ["mozak", "latinexclusive", "latin exclusive", "mrv", "direcional", "tenda", "cury", "cyrela", "even", "gafisa","rossi", "trisul", "eztec", "tecnisa", "brookfield", "plaenge", "mitre", "viver", "rodobens", "patrimar"],
            "name_must_contain_extra_any": [],
            "segment_name_any": ["construtora", "incorporadora", "incorporação", "incorporacao", "engenharia", "imobili", "empreend", "imóveis", "imoveis", "stand de vendas", "estande de vendas", "plantão de vendas", "plantao de vendas", "stand", "estande", "plantão", "plantao"],
            "name_exclude": ["cozinha", "cozinhas", "restaurante", "bar", "lanchonete", "padaria", "café", "cafe", "hamburg", "burger", "pizza", "pizzaria", "bistrô", "bistro", "gourmet", "food", "urbanas"],},
    }

    COLOR_MAP = {
        "🛒 Mercados": "blue",
        "🏫 Escolas": "green",
        "🏫 Faculdades/Universidades": "green",
       "✈️ Agências de Viagens" : "cyan",
        "🏥 Hospitais": "red",
        "💊 Farmácias": "purple",
        "🚗 Automotivos":"darkcyan",
        "🍽️ Restaurantes": "darkred",
        "💪 Academias": "darkgreen",
        "🏬 Shoppings": "cadetblue",
        "🚘 Concessionárias": "lightblue",
        "🏗️ Construtoras / Incorporadoras": "orange",
    }
    def _filter_rows_by_cfg(rows, cfg):
        """Filtra linhas retornadas (Google/OSM) com base em regras da categoria.
        Suporta:
          - name_must_contain (qualquer termo)   -> ex.: ["jeep","fiat","ford"]
          - name_must_contain_extra_any (qualquer termo adicional) -> exige também 1 termo extra
          - name_exclude (qualquer termo)
        """
        if not rows:
            return []
        must = [str(x).strip().lower() for x in (cfg.get("name_must_contain") or []) if str(x).strip()]
        must_extra = [str(x).strip().lower() for x in (cfg.get("name_must_contain_extra_any") or []) if str(x).strip()]
        excl = [str(x).strip().lower() for x in (cfg.get("name_exclude") or []) if str(x).strip()]
        seg_any = [str(x).strip().lower() for x in (cfg.get("segment_name_any") or []) if str(x).strip()]
        conditional_requires = cfg.get("conditional_requires") or []
        if not must and not must_extra and not excl:
            return rows

        out = []
        for r in rows:
            nm = str(r.get("Nome", "")).lower()

            if must and not any(t in nm for t in must):
                # fallback genérico do segmento (ex.: 'stand de vendas' em Construtoras)
                if not (seg_any and any(t in nm for t in seg_any)):
                    continue
            if must_extra and not any(t in nm for t in must_extra):
                continue
            if excl and any(t in nm for t in excl):
                continue


            # Regras condicionais:
            # Se o nome contiver algum termo ambíguo, exige também algum termo contextual.
            if conditional_requires:
                ok_cond = True
                for rule in conditional_requires:
                    if_any = [str(x).strip().lower() for x in (rule.get("if_contains_any") or []) if str(x).strip()]
                    req_any = [str(x).strip().lower() for x in (rule.get("require_any") or []) if str(x).strip()]
                    if if_any and any(t in nm for t in if_any):
                        if req_any and not any(t in nm for t in req_any):
                            ok_cond = False
                            break
                if not ok_cond:
                    continue

            out.append(r)
        return out

    # ===============================
    # STATE
    # ===============================
    if "results_df" not in st.session_state:
        st.session_state.results_df = None
    if "origin" not in st.session_state:
        st.session_state.origin = None
    if "last_error" not in st.session_state:
        st.session_state.last_error = None
    if "debug" not in st.session_state:
        st.session_state.debug = []
    if "addr_input" not in st.session_state:
        st.session_state.addr_input = ""
    if "addr_hist_sel" not in st.session_state:
        st.session_state.addr_hist_sel = "(digitar novo)"
    if "addr_suggestions" not in st.session_state:
        st.session_state.addr_suggestions = []
    if "addr_last_query" not in st.session_state:
        st.session_state.addr_last_query = ""
    if "trigger_suggest" not in st.session_state:
        st.session_state.trigger_suggest = False
    if "addr_input_pending" not in st.session_state:
        st.session_state.addr_input_pending = ""

    _pending_addr = (st.session_state.get("addr_input_pending") or "").strip()
    if _pending_addr:
        st.session_state.addr_input = _pending_addr
        st.session_state.addr_input_pending = ""

    def clear_all():
        st.session_state.results_df = None
        st.session_state.origin = None
        st.session_state.last_error = None
        st.session_state.debug = []
        st.session_state.addr_input = ""
        st.session_state.addr_hist_sel = "(digitar novo)"
        st.session_state.addr_suggestions = []
        st.session_state.addr_last_query = ""
        st.session_state.trigger_suggest = False
        st.session_state.addr_input_pending = ""

    def _on_hist_change():
        sel = (st.session_state.get("addr_hist_sel") or "(digitar novo)").strip()
        if sel and sel != "(digitar novo)":
            st.session_state["addr_input"] = sel

    def _on_addr_change():
        typed = (st.session_state.get("addr_input") or "").strip()
        sel = (st.session_state.get("addr_hist_sel") or "(digitar novo)").strip()
        if sel != "(digitar novo)" and typed != sel:
            st.session_state["addr_hist_sel"] = "(digitar novo)"
        # ENTER/blur no campo agora apenas atualiza as sugestões; a busca fica explícita no botão Buscar
        st.session_state.trigger_suggest = bool(typed)

    # ===============================
    # UTILS
    # ===============================
    def haversine_km(lat1, lon1, lat2, lon2) -> float:
        R = 6371.0088
        phi1 = math.radians(float(lat1))
        phi2 = math.radians(float(lat2))
        dphi = math.radians(float(lat2) - float(lat1))
        dl = math.radians(float(lon2) - float(lon1))
        a = math.sin(dphi/2)**2 + math.cos(phi1)*math.cos(phi2)*math.sin(dl/2)**2
        return 2 * R * math.asin(math.sqrt(a))

    def format_endereco_origem(full: str) -> str:
        """
        Normaliza o "Endereço origem" para ficar somente:
        Rua, Bairro, Cidade, CEP (quando disponível).

        Ex.: "Zona Sul, Rua Teixeira de Melo, Ipanema, Rio de Janeiro, ... , 22410-001, Brasil"
        -> "Rua Teixeira de Melo, Ipanema, Rio de Janeiro, 22410-001"
        """
        if not full:
            return ""

        parts = [p.strip() for p in str(full).split(",") if str(p).strip()]

        # CEP
        cep = None
        for p in parts:
            mm = re.search(r"\b\d{5}-\d{3}\b", p)
            if mm:
                cep = mm.group(0)
                break
        if not cep:
            for p in parts:
                mm = re.search(r"\b\d{8}\b", p)
                if mm:
                    raw = mm.group(0)
                    cep = f"{raw[:5]}-{raw[5:]}"
                    break

        skip_words = ["região", "metropolitana", "geográfica", "intermediária", "imediata", "brasil", "estado"]

        def is_skip(p: str) -> bool:
            pl = str(p).strip().lower()
            if not pl:
                return True
            if pl.startswith("zona "):
                return True
            return any(w in pl for w in skip_words)

        street_keywords = [
            "rua", "avenida", "av.", "av ", "travessa", "estrada", "rodovia",
            "alameda", "praça", "praca", "largo", "beco", "r.", "rod "
        ]

        # Rua
        rua = None
        for p in parts:
            if is_skip(p):
                continue
            pl = p.lower()
            if any(k in pl for k in street_keywords):
                rua = p
                break
        if not rua:
            for p in parts:
                if not is_skip(p):
                    rua = p
                    break

        # Bairro
        bairro = None
        start_idx = parts.index(rua) + 1 if (rua and rua in parts) else 0
        for p in parts[start_idx:]:
            if is_skip(p):
                continue
            if cep and cep in p:
                continue
            if re.fullmatch(r"\d{5}-\d{3}", p) or re.fullmatch(r"\d{8}", p):
                continue
            if rua and p == rua:
                continue
            bairro = p
            break

        # Cidade
        cidade = None
        if bairro and bairro in parts:
            start2 = parts.index(bairro) + 1
        elif rua and rua in parts:
            start2 = parts.index(rua) + 1
        else:
            start2 = 0

        for p in parts[start2:]:
            if is_skip(p):
                continue
            if cep and cep in p:
                continue
            if rua and p == rua:
                continue
            if bairro and p == bairro:
                continue
            if re.fullmatch(r"\d{5}-\d{3}", p) or re.fullmatch(r"\d{8}", p):
                continue
            cidade = p
            break

        out = []
        for x in (rua, bairro, cidade, cep):
            if x and x not in out:
                out.append(x)
        return ", ".join(out)



    def _apply_radius_filter(df: pd.DataFrame, radius_km: float, eps_km: float = 0.15):
        """
        Garante que NADA fora do raio apareça na lista final.
        eps_km evita cortar itens na borda por arredondamento/pequenas diferenças.
        Retorna (df_filtrado, removidos).
        """
        if df is None or df.empty:
            return df, 0
        limit = float(radius_km) + float(eps_km)
        before = len(df)
        df2 = df[df["Distância (km)"] <= limit].copy()
        removed = before - len(df2)
        return df2, removed


    @st.cache_data(ttl=3600)
    def _geocode_one(addr: str):
        g = Nominatim(user_agent=f"prosearch_{int(time.time())}")
        try:
            l = g.geocode(addr, timeout=10)
            if l:
                return l.latitude, l.longitude, l.address
        except GeocoderTimedOut:
            return None, None, None
        except Exception:
            return None, None, None
        return None, None, None

    def geocode_robusto(addr: str):
        if not addr or not addr.strip():
            return None, None, None
        base = addr.strip()
        tries = [base]
        if "brasil" not in base.lower() and "brazil" not in base.lower():
            tries += [f"{base}, Brasil", f"{base}, Brazil"]
        seen=set(); final=[]
        for t in tries:
            if t not in seen:
                seen.add(t); final.append(t)
        for t in final:
            lat, lon, full = _geocode_one(t)
            if lat:
                return float(lat), float(lon), full
        return None, None, None

    # ===============================
    # GOOGLE PLACES (HTTP Endpoints)
    # ===============================
    @st.cache_data(ttl=900)
    def _google_nearby(lat, lon, radius_km, place_type, api_key, keyword=None):
        url = "https://maps.googleapis.com/maps/api/place/nearbysearch/json"
        params = {
            "location": f"{lat},{lon}",
            "radius": int(radius_km * 1000),
            "type": place_type,
            "language": "pt-BR",
            "region": "br",
            "key": api_key,
        }
        if keyword and keyword.strip():
            params["keyword"] = keyword.strip()

        r = requests.get(url, params=params, timeout=15)
        try:
            j = r.json()
        except Exception:
            j = {"status": "BAD_JSON", "error_message": r.text[:200]}
        return j, r.status_code

    @st.cache_data(ttl=24*3600)
    def _google_details(place_id, api_key):
        url = "https://maps.googleapis.com/maps/api/place/details/json"
        params = {
            "place_id": place_id,
            "fields": "formatted_phone_number,international_phone_number,website,rating,user_ratings_total,url,opening_hours,business_status,name,formatted_address",
            "language": "pt-BR",
            "region": "br",
            "key": api_key,
        }
        r = requests.get(url, params=params, timeout=15)
        try:
            j = r.json()
        except Exception:
            j = {"status": "BAD_JSON", "error_message": r.text[:200]}
        return j, r.status_code

    @st.cache_data(ttl=24*3600)
    def fetch_place_photo_bytes(photo_reference: str, api_key: str, maxwidth: int = 520):
        """
        Baixa uma miniatura retangular do local via Places Photo.
        Retorna bytes da imagem (ou None).
        """
        if not photo_reference or not api_key:
            return None
        url = "https://maps.googleapis.com/maps/api/place/photo"
        params = {"maxwidth": int(maxwidth), "photo_reference": photo_reference, "key": api_key}
        try:
            r = requests.get(url, params=params, timeout=15, allow_redirects=True)
            if r.status_code == 200 and r.content:
                return r.content
        except Exception:
            return None
        return None


    def _photo_pick_from_result(r: dict):
        """Extrai (photo_reference, html_attribution) do primeiro item de photos, se existir."""
        photos = (r or {}).get("photos") or []
        if not photos:
            return "", ""
        p0 = photos[0] or {}
        pref = p0.get("photo_reference") or ""
        attrs = p0.get("html_attributions") or []
        attr = attrs[0] if attrs else ""
        return pref, attr

    def google_search_category(lat, lon, radius_km, category_name, api_key, keyword=None, debug=False):
        place_type_cfg = CATEGORIES_CONFIG[category_name].get("google_type")
        # Permite múltiplos types (ex.: incorporadoras também)
        if isinstance(place_type_cfg, (list, tuple, set)):
            place_types = [x for x in place_type_cfg if x]
        else:
            place_types = [place_type_cfg] if place_type_cfg else [None]

        out = []
        seen_pid = set()

        allowed_types = set([t for t in place_types if isinstance(t, str) and t.strip()])

        for place_type in place_types:
            j, http = _google_nearby(lat, lon, radius_km, place_type, api_key, keyword=keyword)

            if debug:
                st.session_state.debug.append({
                    "endpoint": "nearbysearch",
                    "category": category_name,
                    "http_status": http,
                    "status": j.get("status"),
                    "error_message": j.get("error_message"),
                    "results_count": len(j.get("results") or []),
                    "type": place_type,
                    "keyword": keyword or ""
                })

            if j.get("status") != "OK":
                continue

            for r in (j.get("results") or [])[:30]:
                pid = r.get("place_id")
                if pid and pid in seen_pid:
                    continue
                geom = (r.get("geometry") or {}).get("location") or {}
                plat = geom.get("lat")
                plon = geom.get("lng")
                if plat is None or plon is None:
                    continue

                # Filtro genérico por SEGMENTO:
                # Garante que o lugar retornado realmente pertença ao(s) type(s) da categoria.
                # Evita ambiguidades (ex.: "Tenda Cozinhas" cair em Construtoras) sem hardcode por marca.
                if allowed_types:
                    r_types = set([str(x) for x in (r.get("types") or []) if x])
                    if r_types and r_types.isdisjoint(allowed_types):
                        continue

                photo_ref, photo_attr = _photo_pick_from_result(r)

                details={}
                if pid:
                    dj, dh = _google_details(pid, api_key)
                    if debug:
                        res = dj.get("result") or {}
                        st.session_state.debug.append({
                            "endpoint": "details",
                            "category": category_name,
                            "http_status": dh,
                            "status": dj.get("status"),
                            "error_message": dj.get("error_message"),
                            "place_id": pid,
                            "has_phone": bool(res.get("formatted_phone_number") or res.get("international_phone_number")),
                            "has_website": bool(res.get("website"))
                        })
                    if dj.get("status") == "OK":
                        details = dj.get("result") or {}

                tel = details.get("formatted_phone_number") or details.get("international_phone_number") or "-"
                site = details.get("website") or "-"
                rating = details.get("rating", r.get("rating", "-"))
                reviews = details.get("user_ratings_total", r.get("user_ratings_total", "-"))
                maps_url = details.get("url") or (f"https://www.google.com/maps/place/?q=place_id:{pid}" if pid else f"https://www.google.com/maps?q={plat},{plon}")

                status = details.get("business_status") or r.get("business_status") or "Desconhecido"
                horario = "Não disponível"
                oh = details.get("opening_hours") or {}
                if isinstance(oh, dict):
                    wt = oh.get("weekday_text") or []
                    if wt:
                        wd = time.localtime().tm_wday  # 0=segunda
                        if len(wt) == 7:
                            horario = wt[wd]
                        else:
                            horario = "; ".join(wt[:2])
                    if oh.get("open_now") is True:
                        status = "Aberto"
                    elif oh.get("open_now") is False:
                        status = "Fechado"

                out.append({
                    "Nome": r.get("name") or details.get("name") or "Sem nome",
                    "Categoria": category_name,
                    "Endereço": details.get("formatted_address") or r.get("vicinity") or "Endereço não disponível",
                    "Telefone": tel,
                    "Website": site,
                    "Avaliação": rating if rating is not None else "-",
                    "Total Avaliações": reviews if reviews is not None else "-",
                    "Status": status,
                    "Horário": horario,
                    "Latitude": float(plat),
                    "Longitude": float(plon),
                    "Fonte": "Google Places",
                    "Maps": maps_url,
                    "PhotoRef": photo_ref,
                    "PhotoAttr": photo_attr
                })
                if pid:
                    seen_pid.add(pid)

        return out

        if debug:
            st.session_state.debug.append({
                "endpoint": "nearbysearch",
                "category": category_name,
                "http_status": http,
                "status": j.get("status"),
                "error_message": j.get("error_message"),
                "results_count": len(j.get("results") or []),
                "type": place_type,
                "keyword": keyword or ""
            })

        if j.get("status") != "OK":
            return []

        out=[]
        for r in (j.get("results") or [])[:30]:
            pid = r.get("place_id")
            geom = (r.get("geometry") or {}).get("location") or {}
            plat = geom.get("lat")
            plon = geom.get("lng")
            if plat is None or plon is None:
                continue

            photo_ref, photo_attr = _photo_pick_from_result(r)

            details={}
            if pid:
                dj, dh = _google_details(pid, api_key)
                if debug:
                    res = dj.get("result") or {}
                    st.session_state.debug.append({
                        "endpoint": "details",
                        "category": category_name,
                        "http_status": dh,
                        "status": dj.get("status"),
                        "error_message": dj.get("error_message"),
                        "place_id": pid,
                        "has_phone": bool(res.get("formatted_phone_number") or res.get("international_phone_number")),
                        "has_website": bool(res.get("website"))
                    })
                if dj.get("status") == "OK":
                    details = dj.get("result") or {}

            tel = details.get("formatted_phone_number") or details.get("international_phone_number") or "-"
            site = details.get("website") or "-"
            rating = details.get("rating", r.get("rating", "-"))
            reviews = details.get("user_ratings_total", r.get("user_ratings_total", "-"))
            maps_url = details.get("url") or (f"https://www.google.com/maps/place/?q=place_id:{pid}" if pid else f"https://www.google.com/maps?q={plat},{plon}")

            status = details.get("business_status") or r.get("business_status") or "Desconhecido"
            horario = "Não disponível"
            oh = details.get("opening_hours") or {}
            if isinstance(oh, dict):
                wt = oh.get("weekday_text") or []
                if wt:
                    wd = time.localtime().tm_wday  # 0=segunda
                    if len(wt) == 7:
                        horario = wt[wd]
                    else:
                        horario = "; ".join(wt[:2])
                if oh.get("open_now") is True:
                    status = "Aberto"
                elif oh.get("open_now") is False:
                    status = "Fechado"

            out.append({
                "Nome": r.get("name") or details.get("name") or "Sem nome",
                "Categoria": category_name,
                "Endereço": details.get("formatted_address") or r.get("vicinity") or "Endereço não disponível",
                "Telefone": tel,
                "Website": site,
                "Avaliação": rating if rating is not None else "-",
                "Total Avaliações": reviews if reviews is not None else "-",
                "Status": status,
                "Horário": horario,
                "Latitude": float(plat),
                "Longitude": float(plon),
                "Fonte": "Google",
                "Maps": maps_url,
                "PhotoRef": photo_ref,
                "PhotoAttr": photo_attr
            })
        return out

    # ===============================
    # OSM/OVERPASS (cache + fallback)
    # ===============================
    OVERPASS_ENDPOINTS = [
        "https://overpass-api.de/api/interpreter",
        "https://overpass.kumi.systems/api/interpreter",
        "https://overpass.nchc.org.tw/api/interpreter",
    ]

    @st.cache_data(ttl=600)
    def _overpass_call(endpoint: str, q: str):
        r = requests.get(endpoint, params={"data": q}, timeout=45, headers={"User-Agent":"ProSearch/4.0 (Streamlit)"})
        return r.status_code, r.text

    def overpass_fallback(q: str):
        last = None
        for ep in OVERPASS_ENDPOINTS:
            try:
                status, text = _overpass_call(ep, q)
                if status != 200:
                    last = f"{ep} status={status}"
                    continue
                try:
                    return json.loads(text), ep, None
                except Exception as e:
                    last = f"{ep} json error: {e}"
                    continue
            except Exception as e:
                last = f"{ep} exception: {e}"
                continue
        return None, None, last

    def osm_search_category(lat, lon, radius_km, category_name, debug=False):
        r_m = int(radius_km * 1000)
        tags = CATEGORIES_CONFIG[category_name]["osm"]

        tag_queries=[]
        for t in tags:
            k,v = t.split("=")
            tag_queries.append(f'node["{k}"="{v}"](around:{r_m},{lat},{lon});')
            tag_queries.append(f'way["{k}"="{v}"](around:{r_m},{lat},{lon});')
            tag_queries.append(f'relation["{k}"="{v}"](around:{r_m},{lat},{lon});')

        q = f"""
        [out:json][timeout:30];
        (
          {' '.join(tag_queries)}
        );
        out center;
        """
        data, used, err = overpass_fallback(q)

        if debug:
            st.session_state.debug.append({
                "endpoint":"overpass",
                "category": category_name,
                "used": used,
                "error": err,
                "elements": (len((data or {}).get("elements") or []) if data else 0)
            })

        if data is None:
            return []

        out=[]
        for el in data.get("elements", []):
            t = el.get("tags", {}) or {}
            name = t.get("name") or t.get("operator")
            if not name:
                continue
            if "lat" in el and "lon" in el:
                plat, plon = el.get("lat"), el.get("lon")
            else:
                c = el.get("center", {}) or {}
                plat, plon = c.get("lat"), c.get("lon")
            if plat is None or plon is None:
                continue

            street = (t.get("addr:street") or "").strip()
            num = (t.get("addr:housenumber") or "").strip()
            addr = (f"{street}, {num}".strip().strip(",")).strip()
            # Complementos (quando existirem no OSM)
            comp = (t.get("addr:unit") or t.get("addr:flats") or t.get("addr:door") or t.get("addr:floor") or "").strip()
            if comp:
                addr = f"{addr} - {comp}".strip()

            bairro_osm = (t.get("addr:suburb") or t.get("addr:neighbourhood") or t.get("addr:district") or "").strip()
            cidade_osm = (t.get("addr:city") or t.get("addr:municipality") or "").strip()
            cep_osm = (t.get("addr:postcode") or "").strip()
            if not addr:
                addr = "Endereço não informado"

            phone = t.get("phone") or t.get("contact:phone") or "-"
            site = t.get("website") or t.get("contact:website") or "-"
            email = t.get("email") or t.get("contact:email") or ""
            resp = t.get("contact:name") or t.get("contact:person") or ""
            whatsapp = t.get("contact:whatsapp") or t.get("whatsapp") or ""
            instagram = t.get("contact:instagram") or t.get("instagram") or ""
            facebook = t.get("contact:facebook") or t.get("facebook") or ""
            linkedin = t.get("contact:linkedin") or t.get("linkedin") or ""

            out.append({
                "Nome": name,
                "Categoria": category_name,
                "Endereço": addr,
                "Bairro": bairro_osm,
                "Cidade": cidade_osm,
                "CEP": cep_osm,
                "Telefone": phone if phone else "-",
                "Website": site if site else "-",
                "E-mail": email.strip(),
                "Responsável": resp.strip(),
                "WhatsApp": whatsapp.strip(),
                "Instagram": instagram.strip(),
                "Facebook": facebook.strip(),
                "LinkedIn": linkedin.strip(),
                "Avaliação": "-",
                "Total Avaliações": "-",
                "Status": "-",
                "Horário": "-",
                "Latitude": float(plat),
                "Longitude": float(plon),
                "Fonte": "OpenStreetMap",
                "Maps": f"https://www.google.com/maps?q={float(plat)},{float(plon)}",
                "PhotoRef": "",
                "PhotoAttr": ""
            })
        return out

    # ===============================
    # RENDER: CARDS (cards primeiro)
    # ===============================
    def render_cards_grid(df: pd.DataFrame, cols_per_row: int, api_key: str):
        cols_per_row = max(1, min(5, int(cols_per_row)))
        st.markdown("### Resultados")
        if df is None or df.empty:
            st.info("Sem resultados para exibir.")
            return

        rows = df.to_dict("records")
        for start in range(0, len(rows), cols_per_row):
            cols = st.columns(cols_per_row, gap="medium")
            chunk = rows[start:start+cols_per_row]
            for col, r in zip(cols, chunk):
                with col:
                    nome = str(r.get("Nome","Sem nome"))
                    cat = str(r.get("Categoria",""))
                    end = str(r.get("Endereço","Endereço não disponível"))
                    tel = str(r.get("Telefone","-"))
                    web = str(r.get("Website","-"))
                    fonte = str(r.get("Fonte","-"))
                    dist = r.get("Distância (km)", None)
                    dist_str = f"{dist:.2f} km" if isinstance(dist, (float,int)) else "-"

                    rating = r.get("Avaliação","-")
                    nrev = r.get("Total Avaliações","-")
                    maps = r.get("Maps") or f"https://www.google.com/maps?q={r.get('Latitude')},{r.get('Longitude')}"
                    directions = f"https://www.google.com/maps/dir/?api=1&destination={r.get('Latitude')},{r.get('Longitude')}"

                    # Miniatura (retangular) via Places Photo, quando disponível
                    photo_ref = str(r.get("PhotoRef","") or "").strip()
                    photo_attr = str(r.get("PhotoAttr","") or "").strip()
                    img_bytes = None
                    if photo_ref and api_key and api_key.strip():
                        img_bytes = fetch_place_photo_bytes(photo_ref, api_key.strip(), maxwidth=720)

                    if img_bytes:
                        # Imagem com altura fixa para padronizar os cards (150px)
                        try:
                            b64 = base64.b64encode(img_bytes).decode("utf-8")
                            st.markdown(
                                f"""<img src="data:image/jpeg;base64,{b64}"
                                         style="width:100%; height:150px; object-fit:cover; border-radius:16px;"
                                         loading="lazy" />""",
                                unsafe_allow_html=True,
                            )
                        except Exception:
                            st.image(img_bytes, use_container_width=True)
                    else:
                        st.markdown(
                            """<div style="width:100%; height:150px; border-radius:16px;
                                     background: rgba(255,255,255,0.06);
                                     border:1px solid rgba(255,255,255,0.08);
                                     display:flex; align-items:center; justify-content:center; opacity:0.85;">
                                   🖼️ Sem foto
                                 </div>""",
                            unsafe_allow_html=True,
                        )

                    if photo_attr:
                            # Atribuição vem em HTML (exigência do Google). Exibe discreto.
                            st.markdown(f"<div class='small'>📷 {photo_attr}</div>", unsafe_allow_html=True)
                    else:
                        # mantém um espaço visual para consistência
                        st.markdown("<div class='small'>🖼️ (sem foto)</div>", unsafe_allow_html=True)

                    st.markdown(f"""
                    <div class="card">
                      <h4>{nome}</h4>
                      <div class="meta">
                        <span class="badge">{cat}</span>
                        <span class="badge">📏 {dist_str}</span>
                        <span class="badge">🗺️ {fonte}</span>
                      </div>
                      <div class="small">📍 {end}</div>
                      <div class="small">📞 {tel if tel and tel != "-" else "(não informado)"}</div>
                      <div class="small">🌐 {web if web and web != "-" else "(não informado)"}</div>
                      <div class="small">{("⭐ " + str(rating) + " • " + str(nrev) + " avaliações") if str(rating) not in ["-","None","nan"] else ""}</div>
                    </div>
                    """, unsafe_allow_html=True)

                    b1, b2 = st.columns(2)
                    with b1:
                        st.link_button("📍 Maps", maps, use_container_width=True)
                    with b2:
                        st.link_button("🧭 Rotas", directions, use_container_width=True)

    # ===============================
    # RENDER: MAPA INTERATIVO (OSM/Folium)
    # ===============================
    def render_map_interativo(origin, df: pd.DataFrame, zoom: int):
        st.markdown("### Mapa interativo")
        if df is None or df.empty:
            st.info("Sem resultados para mapear.")
            return

        m = folium.Map(location=[origin["lat"], origin["lon"]], zoom_start=max(10, min(18, int(zoom))))
        folium.Marker([origin["lat"], origin["lon"]], tooltip="Você", icon=folium.Icon(color="red")).add_to(m)
        cl = MarkerCluster().add_to(m)

        for _, r in df.iterrows():
            latp, lonp = float(r["Latitude"]), float(r["Longitude"])
            maps = r.get("Maps") or f"https://www.google.com/maps?q={latp},{lonp}"
            directions = f"https://www.google.com/maps/dir/?api=1&destination={latp},{lonp}"
            popup = f"<b>{r.get('Nome','')}</b><br>{r.get('Categoria','')}<br>{r.get('Endereço','')}<br><a href='{maps}' target='_blank'>Maps</a> | <a href='{directions}' target='_blank'>Rotas</a>"
            folium.Marker([latp, lonp], popup=popup, tooltip=r.get("Nome",""), icon=folium.Icon(color=COLOR_MAP.get(r.get("Categoria",""), "blue"))).add_to(cl)

        st_folium(m, width="100%", height=580)

    # ===============================
    # HEADER
    # ===============================
    st.markdown("""
    <div class="header-wrap">
      <div>
        <div class="h1">Mapa Comercial por categorias🗺️</div>
        <div class="sub">Desenvolvido por Sacramento • © V3.20260201</div>
      </div>
      <div class="pill">v4 • 1 arquivo</div>
    </div>
    """, unsafe_allow_html=True)

    # ===============================
    # SIDEBAR
    # ===============================
    with st.sidebar:
        st.header("⚙️ Configurações")

                # ===============================
        # CAMPOS DE BUSCA (sem st.form)
        # - permite sugestões em tempo real no Endereço
        # - a busca só roda quando clicar em "Buscar"
        # ===============================

        default_api_key = "AIzaSyByLLGY4KW3u1kDYmh-puyMwmLsLiTq4H0"
        if not default_api_key:
            default_api_key = os.getenv("GOOGLE_API_KEY", "")
        api_key = default_api_key
        _hist = load_address_history(limit=25)
        _hist_opts = ["(digitar novo)"] + _hist
        _current_hist = st.session_state.get("addr_hist_sel", "(digitar novo)")
        if _current_hist not in _hist_opts:
            _current_hist = "(digitar novo)"
            st.session_state["addr_hist_sel"] = _current_hist
        _hist_index = _hist_opts.index(_current_hist)
        st.selectbox(
            "🕘 Histórico de endereços",
            options=_hist_opts,
            index=_hist_index,
            key="addr_hist_sel",
            on_change=_on_hist_change,
        )

        # Campo principal + botão explícito para atualizar sugestões
        addr_col, sug_col = st.columns([0.74, 0.26], gap="small")
        with addr_col:
            st.text_input(
                "📍 Endereço",
                key="addr_input",
                placeholder="Digite (rua, número, bairro, cidade)...",
                on_change=_on_addr_change,
            )
        with sug_col:
            st.write("")
            st.write("")
            suggest_now = st.button("📌 Sugerir", use_container_width=True)

        # Sugestões estilo Google
        # Observação: no Streamlit o valor do text_input só é consolidado ao confirmar/blur.
        # Então o botão Sugerir torna o fluxo estável tanto local quanto no Cloud.
        _q = (st.session_state.get("addr_input") or "").strip()
        _prev_q = (st.session_state.get("addr_last_query") or "").strip()
        _need_refresh_sugs = bool(st.session_state.pop("trigger_suggest", False)) or suggest_now or ((_q != _prev_q) and len(_q) >= 4)

        if len(_q) < 4:
            _sugs = []
        elif _need_refresh_sugs or not st.session_state.get("addr_suggestions"):
            _sugs = get_suggestions_filtered(_q, limit=6, api_key=api_key)
        else:
            _sugs = st.session_state.get("addr_suggestions") or []

        st.session_state["addr_last_query"] = _q
        st.session_state["addr_suggestions"] = _sugs

        # CSS (visual “Google-like”)
        st.markdown(
            """
            
<style>
  .addr-sug-row{
    display:flex; gap:10px; align-items:flex-start;
    padding:10px 12px;
    border-radius:16px;
    background: rgba(255,255,255,0.04);
    border: 1px solid rgba(255,255,255,0.08);
    margin: 6px 0;
  }
  .addr-sug-text{ width:100%; }
  .addr-sug-main{
    font-size:0.95rem; line-height:1.25rem;
    display:-webkit-box; -webkit-line-clamp:2; -webkit-box-orient:vertical;
    overflow:hidden;
  }
  .addr-sug-sub{
    margin-top:4px;
    opacity:0.75;
    font-size:0.82rem; line-height:1.1rem;
    display:-webkit-box; -webkit-line-clamp:2; -webkit-box-orient:vertical;
    overflow:hidden;
  }
  .addr-sug-text b{ font-weight: 800; }
</style>

            """,
            unsafe_allow_html=True,
        )

        def _pick_suggestion(s: str):
            # Callback (seguro) — preenche o campo e evita sobrescrita pelo histórico
            st.session_state["addr_input"] = s
            st.session_state["addr_hist_sel"] = "(digitar novo)"

        if _sugs:
            st.markdown("**📌 Sugestões (clique em selecionar):**")
            for i, s in enumerate(_sugs, start=1):
                cols = st.columns([0.28, 0.72], gap="small")
                with cols[0]:
                    st.button(
                        "✅ Selecionar",
                        key=f"sug_pick_{i}",
                        on_click=_pick_suggestion,
                        args=(s,),
                        use_container_width=True,
                    )
                    # aplica classe via css selector do container (Streamlit não permite class direta)
                with cols[1]:
                    main_txt, sub_txt = _split_suggestion(s)
                    main_html = _highlight_match(main_txt, _q)
                    sub_html = _highlight_match(sub_txt, _q) if sub_txt else ""
                    sub_div = f"<div class='addr-sug-sub'>{sub_html}</div>" if sub_txt else ""
                    st.markdown(
                        f"""<div class='addr-sug-row'>
                               <div class='addr-sug-text'>
                                 <div class='addr-sug-main'>{main_html}</div>
                                 {sub_div}
                               </div>
                             </div>""",
                        unsafe_allow_html=True,
                    )

            st.caption("Dica: clique em 📌 Sugerir para carregar opções; ao clicar em Buscar, a 1ª sugestão é usada automaticamente.")

        radius = st.slider("📏 Raio (km)", 0.5, 10.0, 3.0, 0.5)

        categories = st.multiselect(
            "🏷️ Categorias",
            options=list(CATEGORIES_CONFIG.keys()),
            default=["🏫 Escolas", "💊 Farmácias", "🍽️ Restaurantes"]
        )

        keyword = st.text_input("🔎 Palavra-chave (opcional)", value="", placeholder="Ex: Zaccaria, 24h, delivery...")
        prefer_google = True

        cols_per_row = st.slider("🧩 Cards por linha", 1, 5, 5, 1)
        top_n = st.slider("📌 Máx. resultados exibidos", 10, 200, 60, 10)
        st.caption("📝 O Top N afeta **somente a exibição**. A exportação pode enviar **tudo** (opção na aba Tabela).")
        zoom = st.slider("🗺️ Zoom do mapa", 10, 18, 14, 1)

        debug = st.checkbox("🧪 Debug (ver status/erros)", value=False)

        go = st.button("🚀 Buscar", type="primary", use_container_width=True)

        st.button("🧹 Limpar", use_container_width=True, on_click=clear_all)

    # ===============================
    # BUSCA
    # ===============================
    if go:
        st.session_state.last_error = None
        # Timestamp da pesquisa (para exportação)
        if ZoneInfo is not None:
            st.session_state.search_dt = datetime.now(ZoneInfo("America/Sao_Paulo"))
        else:
            st.session_state.search_dt = datetime.now()
        st.session_state.results_df = None
        st.session_state.origin = None
        st.session_state.debug = []

        # Endereço final:
        # - Clique numa sugestão => já preencheu addr_input
        # - ENTER no form => usa automaticamente a 1ª sugestão (se existir)
        _addr_raw = (st.session_state.get("addr_input") or "").strip()
        _sugs_now = st.session_state.get("addr_suggestions") or []
        _last_q = (st.session_state.get("addr_last_query") or "").strip()

        if _addr_raw and _sugs_now and (_addr_raw == _last_q) and (_addr_raw != _sugs_now[0]):
            addr = _sugs_now[0]
            st.session_state["addr_input_pending"] = addr  # aplica no próximo rerun, antes do widget
        else:
            addr = _addr_raw

        if not addr:
            st.session_state.last_error = "Informe um endereço válido."
            st.error(st.session_state.last_error)
            st.stop()

        # salva no histórico (não interfere na lógi
        add_address_to_history(addr, limit=25)

        lat, lon, full = geocode_robusto(addr)
        if not lat:
            st.session_state.last_error = "❌ Endereço não encontrado. Tente incluir rua, número, cidade e UF."
        elif not categories:
            st.session_state.last_error = "⚠️ Selecione ao menos 1 categoria."
        else:
            all_rows = []

            # Google (se preferir e tiver key)
            if prefer_google and api_key.strip():
                for cat in categories:
                    cfg = CATEGORIES_CONFIG.get(cat, {}) or {}
                    kw_list = cfg.get("google_keywords") or []
                    if kw_list:
                        for kw in kw_list:
                            kw = str(kw).strip()
                            comb = f"{kw} {keyword}".strip() if keyword else kw
                            rows = google_search_category(lat, lon, radius, cat, api_key.strip(), keyword=comb, debug=debug)
                            all_rows.extend(_filter_rows_by_cfg(rows, cfg))
                    else:
                        rows = google_search_category(lat, lon, radius, cat, api_key.strip(), keyword=keyword, debug=debug)
                        all_rows.extend(_filter_rows_by_cfg(rows, cfg))
                if not all_rows and debug:
                    st.session_state.debug.append({"note":"Google não retornou resultados; fallback OSM."})

            # OSM fallback
            if not all_rows:
                for cat in categories:
                    cfg = CATEGORIES_CONFIG.get(cat, {}) or {}
                    rows = osm_search_category(lat, lon, radius, cat, debug=debug)
                    all_rows.extend(_filter_rows_by_cfg(rows, cfg))

            if not all_rows:
                st.session_state.last_error = "⚠️ Nenhum resultado encontrado. (Se for Google e estiver tudo 'REQUEST_DENIED', verifique billing/restrições/Places API)."
            else:
                df = pd.DataFrame(all_rows)

                df["Distância (km)"] = df.apply(lambda r: haversine_km(lat, lon, r["Latitude"], r["Longitude"]), axis=1).astype(float).round(2)

                # GARANTIA: filtra por raio (evita itens fora do limite)
                df, removed_outside = _apply_radius_filter(df, radius_km=radius, eps_km=0.15)

                df["__key"] = df["Nome"].astype(str) + "_" + df["Latitude"].astype(str) + "_" + df["Longitude"].astype(str)
                df = df.drop_duplicates(subset=["__key"]).drop(columns=["__key"]).reset_index(drop=True)

                df = df.sort_values("Distância (km)", ascending=True).reset_index(drop=True)

                # Mantém TODOS os resultados para exportação (Excel/Google Sheets)
                st.session_state.results_df_all = df.copy()

                # Para exibição no mapa/tabela, usamos apenas o Top N
                df_view = df.head(int(top_n)).reset_index(drop=True)
                st.session_state.results_df = df_view
                st.session_state.origin = {"lat": lat, "lon": lon, "full": full, "short_full": format_endereco_origem(full), "radius": radius, "categories": categories, "keyword": keyword, "removed_outside": int(removed_outside) if "removed_outside" in locals() else 0}

    # ===============================
    # RENDER (persistente)
    # ===============================
    if st.session_state.last_error:
        st.error(st.session_state.last_error)

    if st.session_state.origin:
        o = st.session_state.origin
        st.success(f"📍 {o['full']}")
        st.caption(f"Categorias: {', '.join(o['categories'])} • Raio: {o['radius']} km" + (f" • Keyword: {o['keyword']}" if o.get("keyword") else ""))

    if st.session_state.results_df is not None:
        df_view = st.session_state.results_df
        df_all = st.session_state.get("results_df_all")

        if df_all is None:

            df_all = df_view
        df = df_view
        o = st.session_state.origin

        # KPIs
        c1, c2, c3, c4 = st.columns(4)
        with c1:
            st.markdown(f'<div class="kpi"><div class="label">Total (todos)</div><div class="value">{len(df_all)}</div></div>', unsafe_allow_html=True)
        with c2:
            n_phone = int((df["Telefone"].astype(str).str.strip() != "-").sum()) if "Telefone" in df.columns else 0
            st.markdown(f'<div class="kpi"><div class="label">Com telefone</div><div class="value">{n_phone}</div></div>', unsafe_allow_html=True)
        with c3:
            n_site = int((df["Website"].astype(str).str.strip() != "-").sum()) if "Website" in df.columns else 0
            st.markdown(f'<div class="kpi"><div class="label">Com website</div><div class="value">{n_site}</div></div>', unsafe_allow_html=True)
        with c4:
            near = float(df["Distância (km)"].min()) if not df.empty else 0.0
            st.markdown(f'<div class="kpi"><div class="label">Mais perto</div><div class="value">{near:.2f} km</div></div>', unsafe_allow_html=True)

        st.caption(f"Exibindo **{len(df_view)}** (Top N). Exportação pode enviar **{len(df_all)}** registros.")

        st.markdown('<div class="hr"></div>', unsafe_allow_html=True)

        # CARDS PRIMEIRO
        render_cards_grid(df, cols_per_row=cols_per_row, api_key=api_key)

        st.markdown('<div class="hr"></div>', unsafe_allow_html=True)

        tab1, tab2 = st.tabs(["🗺️ Mapa", "📊 Tabela"])
        with tab1:
            render_map_interativo(o, df, zoom=zoom)
        with tab2:
            show_cols = ["Nome","Categoria","Distância (km)","Endereço","Telefone","Website","Avaliação","Total Avaliações","Fonte","Maps"]
            show_cols = [c for c in show_cols if c in df.columns]
            st.dataframe(df[show_cols], use_container_width=True, hide_index=True)
            csv = df.to_csv(index=False, encoding="utf-8-sig")
            st.download_button("📥 Baixar CSV", csv, "busca.csv", "text/csv")


            st.markdown("<div class='hr'></div>", unsafe_allow_html=True)
            export_only_topn = st.checkbox("📌 Exportar somente o Top N exibido (senão exporta tudo)", value=False, key="export_only_topn")
            df_export = df_view if export_only_topn else df_all

            st.markdown("### 📤 Exportar resultados na planilha (Excel)")

            st.caption("A exportação é **incremental**: o app encontra a última linha preenchida e adiciona os novos registros na próxima linha.")

            up = st.file_uploader("📎 Envie a planilha Excel (.xlsx) para continuar incremental", type=["xlsx"], key="export_xlsx_uploader")

            template_bytes = None
            template_name = EXPORT_TEMPLATE_FILENAME
            if up is not None:
                template_bytes = up.getvalue()
                template_name = up.name
            else:
                default_fp = os.path.join(_app_base_dir(), EXPORT_TEMPLATE_FILENAME)
                if os.path.exists(default_fp):
                    try:
                        template_bytes = Path(default_fp).read_bytes()
                        st.info(f"Usando planilha padrão encontrada na pasta do app: **{EXPORT_TEMPLATE_FILENAME}**")
                    except Exception as e:
                        st.warning(f"Não consegui ler a planilha padrão ({EXPORT_TEMPLATE_FILENAME}): {e}")
                else:
                    st.warning(f"Envie uma planilha .xlsx ou coloque **{EXPORT_TEMPLATE_FILENAME}** na mesma pasta do app.")

            sheet_name = None
            if template_bytes:
                try:
                    wb_tmp = load_workbook(BytesIO(template_bytes), read_only=False)
                    sheets = wb_tmp.sheetnames
                    default_idx = sheets.index(DEFAULT_EXPORT_SHEET_NAME) if DEFAULT_EXPORT_SHEET_NAME in sheets else 0
                    sheet_name = st.selectbox("📄 Aba de destino", options=sheets, index=default_idx)
                except Exception as e:
                    st.error(f"Erro ao abrir a planilha: {e}")
                    template_bytes = None

            dedup = st.checkbox("🧠 Evitar duplicados (Cliente + Contato)", value=True)

            can_export = template_bytes is not None and sheet_name is not None and (df_export is not None) and (not df_export.empty)
            exp_btn = st.button("📤 Exportar resultados na planilha", type="secondary", use_container_width=True, disabled=not can_export)

            if exp_btn:
                try:
                    dt = st.session_state.get("search_dt")
                    updated_bytes, stats = export_results_incremental_xlsx(
                        df=df_export,
                        template_bytes=template_bytes,
                        sheet_name=sheet_name,
                        executiva="Vanessa",
                        updated_dt=dt,
                        dedup=dedup,
                        google_api_key_for_photo=api_key,
                    )

                    out_name = template_name.replace(".xlsx", "")
                    out_filename = f"{out_name} - atualizado.xlsx"

                    st.success(
                        f"✅ Exportação pronta! Registros adicionados: **{stats['added']}**"
                        + (f" • Duplicados ignorados: **{stats['skipped']}**" if dedup else "")
                    )
                    st.download_button(
                        "⬇️ Baixar planilha atualizada (.xlsx)",
                        data=updated_bytes,
                        file_name=out_filename,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True,
                    )


                    # 🔎 Prévia rápida do que foi gravado (últimas linhas)
                    try:
                        wb_prev = load_workbook(BytesIO(updated_bytes))
                        ws_prev = wb_prev[stats.get("sheet") or sheet_name]
                        header_row_prev, cols_prev = _ensure_export_layout(ws_prev)
                        last_prev = _last_filled_row_any(
                            ws_prev, header_row_prev,
                            [cols_prev["Segmento"], cols_prev["Executiva"], cols_prev["Cliente"], cols_prev["Contato"]]
                        )
                        first_prev = max(header_row_prev + 1, last_prev - 15)
                        preview_rows = []
                        for rr in range(first_prev, last_prev + 1):
                            preview_rows.append({
                                "Segmento": ws_prev.cell(rr, cols_prev["Segmento"]).value,
                                "Executiva": ws_prev.cell(rr, cols_prev["Executiva"]).value,
                                "Endereço origem": ws_prev.cell(rr, cols_prev.get("Endereço origem")).value if cols_prev.get("Endereço origem") else None,
                                "Cliente": ws_prev.cell(rr, cols_prev["Cliente"]).value,
                                "Endereço": ws_prev.cell(rr, cols_prev.get("Endereço")).value if cols_prev.get("Endereço") else None,
                                "Contato": ws_prev.cell(rr, cols_prev["Contato"]).value,
                                "Atualizado em": ws_prev.cell(rr, cols_prev["Atualizado em"]).value,
                            })
                        st.markdown("#### 🔎 Prévia (últimas linhas da aba)")
                        st.dataframe(pd.DataFrame(preview_rows), use_container_width=True, hide_index=True)
                    except Exception as _e:
                        st.info("Prévia não disponível (não atrapalha a exportação).")
                except Exception as e:
                    st.error(f"❌ Falha ao exportar: {e}")



            st.markdown("<div class='hr'></div>", unsafe_allow_html=True)
            st.markdown("### 📤 Exportar resultados para Google Planilhas")

            st.caption(
                "A exportação é **incremental**: o app adiciona os novos registros no final da aba. "
                "Campos a partir da coluna C: Segmento_TXT | Executiva Pixel | Empresa (Cliente) | Responsável pela Empresa | Telefone | E-mail | Já fiz contato? | Data de contato | Observações | Site | Endereço | Bairro | CEP | Atualizado em | Foto | Foto_AppSheet. As colunas A e B não são alteradas, e a coluna R recebe apenas a URL da imagem."
            )

            # Planilha fixa (não precisa preencher nada)
            gsheet_url = DEFAULT_GSHEET_ID
            if not (gsheet_url or "").strip():
                st.error("❌ DEFAULT_GSHEET_ID está vazio. Defina a variável de ambiente GSHEET_ID ou ajuste DEFAULT_GSHEET_ID no código.")

            worksheet_name = DEFAULT_GSHEET_TAB

            # Credenciais:
            # - Cloud: detecta automaticamente via _safe_secrets_get("gcp_service_account")
            # - Local: opcionalmente, abra o expander e envie/aponte o JSON
            sa_info_override = None
            cred_path_override = None
            if _safe_secrets_has("gcp_service_account"):
                st.success("✅ Credenciais detectadas via st.secrets (Cloud). Você não precisa informar nada aqui.")
            else:
                with st.expander("🔐 Credenciais (somente se estiver rodando LOCAL)", expanded=False):
                    st.caption("No local, envie o JSON do Service Account ou informe o caminho do arquivo.")
                    sa_up = st.file_uploader("📎 Service Account JSON (opcional)", type=["json"], key="gs_sa_json")
                    if sa_up is not None:
                        try:
                            sa_info_override = json.loads(sa_up.getvalue().decode("utf-8"))
                            st.success("✅ JSON carregado. Vou usar estas credenciais para exportar.")
                        except Exception as _e:
                            st.error(f"JSON inválido: {_e}")
                            sa_info_override = None

                    default_local_path = os.getenv("GOOGLE_APPLICATION_CREDENTIALS", "") or os.path.join(_app_base_dir(), "service_account.json")
                    cred_path_override = st.text_input("📁 Caminho do JSON (opcional)", value=default_local_path, key="gs_sa_path").strip()
                    if cred_path_override and os.path.exists(cred_path_override):
                        st.info("📁 Arquivo encontrado no caminho informado.")

            # 100% automático: sempre evita duplicados
            dedup_gs = True

            can_export_gs = (df_export is not None) and (not df_export.empty) and bool((gsheet_url or "").strip())
            exp_btn_gs = st.button(
                "📤 Exportar para Google Planilhas",
                type="secondary",
                use_container_width=True,
                disabled=not can_export_gs,
            )

            if exp_btn_gs:
                try:
                    st.session_state.gsheet_url = gsheet_url
                    dt = st.session_state.get("search_dt")
                    stats = export_results_incremental_gsheet(
                        df=df_export,
                        spreadsheet_url_or_id=gsheet_url,
                        worksheet_name=worksheet_name,
                        executiva="Vanessa",
                        updated_dt=dt,
                        dedup=dedup_gs,
                        sa_info_override=sa_info_override,
                        cred_path_override=cred_path_override,
                        google_api_key_for_photo=api_key,
                    )

                    st.success(
                        f"✅ Exportação concluída! Registros adicionados: **{stats['added']}**"
                        + (f" • Duplicados ignorados: **{stats['skipped']}**" if dedup_gs else "")
                    )

                    try:
                        url_open = f"https://docs.google.com/spreadsheets/d/{stats['spreadsheet_id']}/edit"
                        st.link_button("🔗 Abrir planilha no Google", url_open, use_container_width=True)
                    except Exception:
                        pass

                except Exception as e:
                    st.error(f"❌ Falha ao exportar para Google Planilhas: {e}")



        if debug and st.session_state.debug:
            with st.expander("🧪 Debug (status/erros do Google/Overpass)", expanded=False):
                st.json(st.session_state.debug)

    else:
        if not st.session_state.last_error:
            st.info("👈 Configure endereço, raio e categorias e clique em **Buscar**. Os resultados ficam na tela.")

    st.caption("Mapa Comercial por categorias")




# ===============================
# ENTRYPOINT
# - VS Code Play (python): abre o Streamlit automaticamente
# - streamlit run: executa normalmente
# ===============================
def _is_running_under_streamlit() -> bool:
    try:
        from streamlit.runtime.scriptrunner import get_script_run_ctx
        return get_script_run_ctx() is not None
    except Exception:
        return False


def _is_port_free(port: int, host: str = "127.0.0.1") -> bool:
    try:
        with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as s:
            s.settimeout(0.25)
            return s.connect_ex((host, int(port))) != 0
    except Exception:
        return False


def _pick_free_port(start: int = 8501, end: int = 8515) -> int:
    for p in range(int(start), int(end) + 1):
        if _is_port_free(p):
            return p
    return int(start)


def _run_streamlit_from_python():
    # Quando você clica "Play" no VS Code, ele roda "python arquivo.py".
    # Aqui nós iniciamos o Streamlit automaticamente.
    port = _pick_free_port(8501, 8515)
    args = [
        sys.executable, "-m", "streamlit", "run", os.path.abspath(__file__),
        "--server.port", str(port),
        "--server.address", "127.0.0.1",
        "--server.headless", "false",
    ]
    # Não usar ngrok aqui; apenas local.
    print("=" * 60)
    print("🚀 Iniciando Streamlit automaticamente (VS Code Play / python)...")
    print(f"🔗 URL: http://127.0.0.1:{port}")
    print("=" * 60)
    # Bloqueia no processo do Streamlit (CTRL+C para parar)
    subprocess.run(args)


if __name__ == "__main__":
    if _is_running_under_streamlit():
        main()
    else:
        _run_streamlit_from_python()
